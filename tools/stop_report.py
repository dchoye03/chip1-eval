"""STOP 전류 보고표 생성 — 팀 테스트 플랜(STOP Procedure) 형식.

results/stop_current.csv (GUI 'Stop Current' 측정 누적)를 읽어
results/STOP_current_report.xlsx 를 만든다:
  - 조건표 (플랜의 Test Items 표 형식: package/SCK/SDA/AINP/AINN/VDD)
  - 시퀀스 기록
  - 시료별 결과 (기존#1~3 / 신규#1~3) + 빈소켓 베이스라인 차감
  - 그룹 평균 + 신규 vs 기존 변화율

같은 라벨이 여러 번 측정됐으면 **마지막 측정**을 쓴다 (재측정 = 갱신).
사용: python tools\\stop_report.py   (측정 추가 후 재실행하면 갱신)
"""
import csv
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import chip1_autotest as g   # RESULTS_DIR, _vendor 부트스트랩 재사용

def csv_path(run: str | None = None) -> Path:
    """런(테스트 그룹)별 CSV — run 없으면 기본 캠페인."""
    return g.RESULTS_DIR / (f"stop_current_{run}.csv" if run
                            else "stop_current.csv")


def out_path(run: str | None = None) -> Path:
    return g.RESULTS_DIR / (f"STOP_current_report_{run}.xlsx" if run
                            else "STOP_current_report.xlsx")

def is_baseline(label: str) -> bool:
    return "baseline" in label.lower() or "빈소켓" in label


def split_groups(rows: dict) -> tuple[list[str], list[str]]:
    """라벨 분류: '기존*' = 기존 그룹 / 베이스라인 제외 나머지 = 신규 그룹.

    신규 시료는 실제 시료 ID(예: 신규#1)로 기록될 수 있으므로 이름을
    강제하지 않는다 — 기존/베이스라인이 아닌 것은 전부 신규로 간주."""
    old = sorted(k for k in rows if k.startswith("기존"))
    new = sorted(k for k in rows
                 if not k.startswith("기존") and not is_baseline(k))
    return old, new


def load_rows(path: Path) -> dict:
    if not path.exists():
        sys.exit(f"측정 CSV 없음: {path} — GUI Stop Current부터 실행")
    latest: dict = {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        rdr = csv.DictReader(f)
        if rdr.fieldnames:                          # BOM/공백 방어
            rdr.fieldnames = [fn.lstrip("﻿").strip()
                              for fn in rdr.fieldnames]
        if not rdr.fieldnames or "label" not in rdr.fieldnames:
            sys.exit("CSV 헤더 손상 — GUI로 측정 1회 실행하면 자동 복구됨")
        for row in rdr:
            lab = (row.get("label") or "").strip()
            # vdd 정규화: "5"/"5.0"/공백 혼재(세션별 표기 차) → "5.0"/"3.3"
            # 둘로 통일. 구버전 CSV(컬럼 없음) = 3.3 취급.
            raw = (row.get("vdd") or "3.3").strip()
            vdd = "5.0" if raw.startswith("5") else "3.3"
            if lab:
                latest[(lab, vdd)] = row            # 같은 (라벨,전압)만 갱신
    return latest


MANUAL_ROWS = ["기존#1", "기존#2", "기존#3",
               "신규#1", "신규#2", "신규#3", "빈소켓_baseline"]
MANUAL_READ_COLS = (3, 7)   # C..G = 판독 1~5
MANUAL_NOTE_COL = 10        # J = 비고

# 수동 시트 2종(3.3V/5V) × 섹션 2종(STOP/Normal). 수동 = 명령/점퍼로 상태를
# 만들고 DMM 판독값을 손으로 기입 (팀 요청: Normal 동작 전류 병기).
MANUAL_SHEETS = {
    "3V3_수동": dict(
        color="1F3864",
        title="STOP Procedure — VDD 3.3V 수동 측정 (자동측정 교차검증)",
        method=("방법: 선A=J2 3.3V핀, A0·A1 분리, 수동 점퍼로 CLK 제어 — "
                "GND에 꽂으면 Normal(동작), DVDD에 꽂으면(캡 있으면 떼기만) "
                "STOP(PD). DMM 판독값을 C~G열에 직접 입력 (µA). "
                "평균/차감 자동. 입력값은 보고표 재생성 시 보존됨."),
    ),
    "5V_수동": dict(
        color="7B3F00",
        title="STOP Procedure — VDD 5.0V 수동 측정 (레벨시프터 확보 전까지)",
        method=("방법: 선A=J2 5V핀 + J206 캡(4.7k 풀업 연결), A0·A1 분리 — "
                "GND에 꽂으면 Normal(동작), 떼면 풀업이 5V로 올려 STOP(PD). "
                "(A0 연결 시 펌웨어식: `stop pd ext`) 판독값 C~G열 직접 입력 — "
                "STOP은 µA, Normal은 mA 단위 (표별 단위 통일, 빈소켓 행 포함). "
                "⚠ 캡 구성에서 Normal(GND) 판독 시 풀업 루프 ~1.06mA 포함 — "
                "캡 빼고 재거나 빈소켓 행 차감(I열)으로 제거."),
    ),
}
MANUAL_SECTIONS = [
    ("stop", "① STOP(PD) 전류 — CLK High 유지, 잠든 상태 (µA)"),
    ("normal", "② Normal(동작) 전류 — CLK Low 유지, 변환 중 (mA)"),
]


_NON_LABEL = {"시료", "기존 평균", "신규 평균"}
_NON_LABEL_PREFIX = ("①", "②", "STOP Procedure", "방법:")


def _load_manual_values(path: Path) -> dict:
    """기존 보고표의 수동 시트들에서 사용자가 입력한 판독값/비고를 회수
    (보고표 재생성 때 보존). 라벨 행은 제외 목록 방식으로 판별하므로
    임의 시료명(새 테스트 그룹)도 그대로 보존된다.
    {(시트, 섹션, 라벨): ([판독5개], 비고)}"""
    saved = {}
    if not path.exists():
        return saved
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        for name in MANUAL_SHEETS:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            section = "stop"
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=2).value
                if v is None:
                    continue
                s = str(v).strip()
                if s.startswith("①"):
                    section = "stop"
                    continue
                if s.startswith("②"):
                    section = "normal"
                    continue
                if s in _NON_LABEL or s.startswith(_NON_LABEL_PREFIX):
                    continue
                reads = [ws.cell(row=r, column=c).value
                         for c in range(MANUAL_READ_COLS[0],
                                        MANUAL_READ_COLS[1] + 1)]
                note = ws.cell(row=r, column=MANUAL_NOTE_COL).value
                if any(x is not None for x in reads) or note:
                    saved[(name, section, s)] = (reads, note)
        wb.close()
    except Exception:                     # noqa: BLE001 - 보존 실패 시 빈 시트
        pass
    return saved


def _build_manual_sheet(wb, name: str, cfg: dict, saved: dict,
                        old_labels: list, new_labels: list):
    """수동 측정 시트 1장 생성: STOP/Normal 두 표, 판독값만 손으로 넣으면
    평균/베이스라인 차감은 수식이 계산. 시료 행은 그룹 목록(가변)에서 생성."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(name)
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor=cfg["color"])
    head_font = Font(color="FFFFFF", bold=True)

    ws.merge_cells("B2:J2")
    t = ws.cell(row=2, column=2, value=cfg["title"])
    t.font = Font(bold=True, size=14, color=cfg["color"])
    ws.cell(row=3, column=2, value=cfg["method"]).font = Font(size=9, italic=True)

    r = 6
    for key, sec_title in MANUAL_SECTIONS:
        unit = "µA" if key == "stop" else "mA"
        hdr = ["시료", "판독1", "판독2", "판독3", "판독4", "판독5",
               f"평균 ({unit})", f"베이스라인 차감 ({unit})", "비고"]
        ws.cell(row=r, column=2, value=sec_title).font = \
            Font(bold=True, size=11, color=cfg["color"])
        for i, h in enumerate(hdr):
            c = ws.cell(row=r + 1, column=2 + i, value=h)
            c.fill = head_fill
            c.font = head_font
            c.border = box
            c.alignment = Alignment(horizontal="center")
        # 행 구성: 기존 그룹(+평균) + 신규 그룹(+평균) + 빈소켓 — 그룹 크기 가변
        grp_fill = PatternFill("solid", fgColor="E8EAF0")
        n_rows = (len(old_labels) + (1 if old_labels else 0)
                  + len(new_labels) + (1 if new_labels else 0) + 1)
        base_row = r + 2 + n_rows - 1               # 각 표의 마지막 = baseline
        rr = r + 2

        def sample_row(label):
            ws.cell(row=rr, column=2, value=label).border = box
            reads, note = saved.get((name, key, label), ([None] * 5, None))
            for j in range(5):
                cell = ws.cell(row=rr, column=3 + j)
                cell.border = box
                cell.number_format = "0.000"
                if j < len(reads) and reads[j] is not None:
                    cell.value = reads[j]
            avg = ws.cell(row=rr, column=8,
                          value=f"=IFERROR(AVERAGE(C{rr}:G{rr}),\"\")")
            avg.border = box
            avg.number_format = "0.000"
            sub = ws.cell(row=rr, column=9)
            if rr != base_row:
                sub.value = f"=IFERROR(H{rr}-$H${base_row},\"\")"
            sub.border = box
            sub.number_format = "0.000"
            nc = ws.cell(row=rr, column=MANUAL_NOTE_COL)
            nc.border = box
            if note:
                nc.value = note

        def avg_row(title, first, last):
            c0 = ws.cell(row=rr, column=2, value=title)
            c0.border = box
            c0.font = Font(bold=True)
            c0.fill = grp_fill
            for c in range(3, 8):
                cell = ws.cell(row=rr, column=c)
                cell.border = box
                cell.fill = grp_fill
            avg = ws.cell(row=rr, column=8,
                          value=f"=IFERROR(AVERAGE(H{first}:H{last}),\"\")")
            avg.border = box
            avg.fill = grp_fill
            avg.font = Font(bold=True)
            avg.number_format = "0.000"
            sub = ws.cell(row=rr, column=9,
                          value=f"=IFERROR(H{rr}-$H${base_row},\"\")")
            sub.border = box
            sub.fill = grp_fill
            sub.font = Font(bold=True)
            sub.number_format = "0.000"
            nc = ws.cell(row=rr, column=MANUAL_NOTE_COL)
            nc.border = box
            nc.fill = grp_fill

        for group, gname in ((old_labels, "기존 평균"),
                             (new_labels, "신규 평균")):
            if not group:
                continue
            first = rr
            for lb in group:
                sample_row(lb)
                rr += 1
            avg_row(gname, first, rr - 1)
            rr += 1
        sample_row("빈소켓_baseline")
        rr += 1
        r = base_row + 3        # 표 사이 여백 2행

    widths = {2: 16, 3: 9, 4: 9, 5: 9, 6: 9, 7: 9, 8: 12, 9: 16, 10: 24}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _manual_mean(saved: dict, sheet: str, sec: str, labels: list) -> float | None:
    """수동 시트 입력값에서 그룹 평균 산출 (없으면 None)."""
    per = []
    for lb in labels:
        entry = saved.get((sheet, sec, lb))
        if entry:
            nums = [v for v in entry[0] if isinstance(v, (int, float))]
            if nums:
                per.append(sum(nums) / len(nums))
    return sum(per) / len(per) if per else None


def _build_summary_sheet(wb, data: dict):
    """맨 앞 '요약' 시트 — 상부 보고용.

    구성 (2026-08-13 팀 지시 반영): 판정 기준 명시(Test Plan 'STOP':
    Shutdown 모드 VDD 누설 < 10µA @ VDD 5.0V) / 신규·기존 그룹 분리
    (신규 = 시험 대상 PASS, 기존 = 비교군 FAIL) / Normal 데이터 병기."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("요약", 0)
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill("solid", fgColor="E2EFDA")
    bad_fill = PatternFill("solid", fgColor="FCE4E4")

    CRIT_UA = 10.0          # Test Plan 'STOP': VDD < 10 µA (@ VDD 5.0V)

    def fu(v, unit="µA"):
        return f"{v:.3f} {unit}" if isinstance(v, float) else "측정 대기"

    ws.merge_cells("B2:G2")
    t = ws.cell(row=2, column=2,
                value="STOP 전류 시험 요약 — 패키지 변경(신규 8SOP) 검증")
    t.font = Font(bold=True, size=14, color="1F3864")

    crit = ws.cell(row=4, column=2, value=(
        "판정 기준: 칩 테스트 플랜 'Test Items - STOP' — "
        "Shutdown(STOP) 모드 누설 VDD < 10 µA (VDD 5.0V, SCK High 유지)"))
    crit.font = Font(bold=True, size=10)

    def section(r0, title, verdict, verdict_fill, rows):
        ws.cell(row=r0, column=2, value=title).font = Font(bold=True, size=12,
                                                           color="1F3864")
        v = ws.cell(row=r0, column=6, value=verdict)
        v.font = Font(bold=True, size=12)
        v.fill = verdict_fill
        v.border = box
        v.alignment = Alignment(horizontal="center")
        hdr = ["조건", "측정 평균 (차감)", "기준", "판정", "비고"]
        for i, h in enumerate(hdr):
            c = ws.cell(row=r0 + 1, column=2 + i, value=h)
            c.fill = head_fill
            c.font = head_font
            c.border = box
            c.alignment = Alignment(horizontal="center")
        for j, row in enumerate(rows):
            rr = r0 + 2 + j
            for i, val in enumerate(row):
                cell = ws.cell(row=rr, column=2 + i)
                if isinstance(val, tuple):
                    cell.value, f = val
                    cell.fill = f
                else:
                    cell.value = val
                cell.border = box
                cell.alignment = Alignment(horizontal="center")
        return r0 + 2 + len(rows)

    new5, old5 = data.get("new_5"), data.get("old_5")
    ratio = (f"기준의 약 {old5 / CRIT_UA:.0f}배 초과 — 비정상 누설"
             if isinstance(old5, float) else "비정상 누설")
    dn = (f"신규 대비 +{data['old_5n'] - data['new_5n']:.2f} mA "
          "(STOP 누설과 동일량 = 상태 무관 고정 누설)"
          if isinstance(data.get("old_5n"), float)
          and isinstance(data.get("new_5n"), float) else "")

    # ---- 신규 (시험 대상) ----
    end = section(
        6, "■ 신규 패키지 (신규#1/B/C) — 시험 대상",
        "판정: PASS", ok_fill,
        [
            ("VDD 5.0V · STOP", fu(new5), "< 10 µA", ("PASS", ok_fill),
             "기준 대비 여유 약 30배"),
            ("VDD 3.3V · STOP", fu(data.get("new_33")), "(참고)", "정상",
             "정식 기준은 5.0V 조건"),
            ("VDD 5.0V · Normal", fu(data.get("new_5n"), "mA"),
             "기준 없음 (참고)", "정상", "동작 전류"),
            ("VDD 3.3V · Normal", fu(data.get("new_33n"), "mA"),
             "기준 없음 (참고)", "정상", ""),
        ])

    # ---- 기존 (비교군) ----
    end = section(
        end + 2, "■ 기존 패키지 (기존#1~3) — 비교군 (종전 패키지 도면)",
        "판정: FAIL", bad_fill,
        [
            ("VDD 5.0V · STOP", fu(old5), "< 10 µA", ("FAIL", bad_fill),
             ratio),
            ("VDD 3.3V · STOP", fu(data.get("old_33")), "(참고)", "충족",
             "결함은 5V에서만 발현"),
            ("VDD 5.0V · Normal", fu(data.get("old_5n"), "mA"),
             "기준 없음 (참고)", "-", dn),
            ("VDD 3.3V · Normal", fu(data.get("old_33n"), "mA"),
             "기준 없음 (참고)", "-", "신규와 동등"),
        ])

    notes = [
        ("결론", "신규 패키지: 기준 충족 (0.32µA < 10µA) — 개선 목적 달성", True),
        ("결론", "기존 패키지: 기준 55배 초과 (553µA) — 불합격", True),
        (None, None, None),
        ("참고", "기존 칩의 누설은 5V에서만 나타남 (3.3V에서는 정상)", False),
        ("참고", "누설량은 Normal 상태에서도 동일 (+0.55mA) — 동작 여부와 무관한 고정 누설", False),
        (None, None, None),
        ("시험", "2026-08-12~13, 시료: 기존 3개 · 신규 3개 — 각 5회 판독 평균", False),
        ("측정", "DM3058E를 VDD 경로에 직렬 연결, 빈 소켓 오프셋 차감 (단위: STOP=µA, Normal=mA)", False),
        ("근거", "칩 테스트 플랜 'STOP' 항목 / 상세 데이터: 3V3_자동 · 3V3_수동 · 5V_수동 시트", False),
    ]
    for i, (tag, text, bold) in enumerate(notes):
        if tag is None:
            continue
        tc = ws.cell(row=end + 2 + i, column=2, value=f"[{tag}]")
        tc.font = Font(bold=True, size=10)
        tc.alignment = Alignment(horizontal="center")
        ws.cell(row=end + 2 + i, column=3, value=text).font = \
            Font(bold=bool(bold), size=10)

    widths = {2: 22, 3: 20, 4: 16, 5: 12, 6: 40, 7: 6}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_auto_sheet(wb, rows, *, sheet_name, title, sck_val, vdd_val,
                      foot, first=False):
    """자동측정(CSV) 시트 1장 — VDD 조건별 (3V3_자동 / 5V_자동).
    반환: (group_avgs, baseline, old_labels, new_labels)"""
    baseline = None
    for k, r in rows.items():
        if is_baseline(k):
            baseline = float(r["avg_uA"])
    OLD_LABELS, NEW_LABELS = split_groups(rows)

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if first:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)

    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(color="FFFFFF", bold=True)
    grp_fill = PatternFill("solid", fgColor="D9E2F2")

    def put(r, c, v, bold=False, fill=None, fmt=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = box
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        return cell

    # ---- 제목 ----
    ws.merge_cells("B2:J2")
    t = ws.cell(row=2, column=2, value=title)
    t.font = Font(bold=True, size=14, color="1F3864")

    # ---- 조건표 (플랜 형식) ----
    cond_hdr = ["package", "SCK", "SDA", "AINP", "AINN", "SCL Freq.", "VDD"]
    cond_val = ["SOP-8", sck_val, "open", "open", "open", "-", vdd_val]
    for i, (h, v) in enumerate(zip(cond_hdr, cond_val)):
        put(4, 2 + i, h, fill=head_fill).font = head_font
        put(5, 2 + i, v)
    ws.cell(row=6, column=2, value=foot)
    ws.cell(row=6, column=2).font = Font(size=9, italic=True)

    # ---- 시퀀스 ----
    seq = ["# power on", "# set SCK '0'",
           "# wait 2 ms  (resetb deassert 1.5 ms + eFuse cloning 0.5 ms)",
           "# set SCK '1'", "# wait 500 us  (power down > 200 us)",
           "# measure current of VDD (DM3058E, DVDD 직렬)"]
    ws.cell(row=8, column=2, value="Sequence").font = Font(bold=True)
    for i, s in enumerate(seq):
        ws.cell(row=9 + i, column=2, value=s).font = Font(size=9)

    # ---- 결과표 ----
    r0 = 16
    hdr = ["시료", "n", "측정 평균 (µA)", "min", "max",
           "베이스라인 차감 (µA)", "비고"]
    for i, h in enumerate(hdr):
        put(r0, 2 + i, h, fill=head_fill).font = head_font

    def sample_row(r, label):
        d = rows.get(label)
        put(r, 2, label)
        if d is None:
            put(r, 3, "-"); put(r, 4, "미측정"); put(r, 5, "-"); put(r, 6, "-")
            put(r, 7, "-"); put(r, 8, "")
            return None
        avg = float(d["avg_uA"])
        put(r, 3, int(d["n"]))
        put(r, 4, avg, fmt="0.000")
        put(r, 5, float(d["min_uA"]), fmt="0.000")
        put(r, 6, float(d["max_uA"]), fmt="0.000")
        if baseline is not None:
            put(r, 7, avg - baseline, fmt="0.000")
        else:
            put(r, 7, "-")
        put(r, 8, d["datetime"][:16])
        return avg

    r = r0 + 1
    group_avgs = {}
    for name, labels in (("기존", OLD_LABELS), ("신규", NEW_LABELS)):
        vals = []
        for lb in labels:
            v = sample_row(r, lb)
            if v is not None:
                vals.append(v)
            r += 1
        put(r, 2, f"{name} 평균", bold=True, fill=grp_fill)
        if vals:
            m = sum(vals) / len(vals)
            group_avgs[name] = m
            put(r, 3, len(vals), fill=grp_fill)
            put(r, 4, m, fill=grp_fill, fmt="0.000").font = Font(bold=True)
            if baseline is not None:
                put(r, 7, m - baseline, fill=grp_fill, fmt="0.000").font = Font(bold=True)
            else:
                put(r, 7, "-", fill=grp_fill)
            put(r, 5, "-", fill=grp_fill); put(r, 6, "-", fill=grp_fill)
            put(r, 8, "", fill=grp_fill)
        else:
            for c in range(3, 9):
                put(r, c, "-", fill=grp_fill)
        r += 1

    # 베이스라인 행
    put(r, 2, "빈소켓 baseline", bold=True)
    if baseline is not None:
        put(r, 4, baseline, fmt="0.000")
        put(r, 8, "전원 LED + 납땜 기준칩 + 보드 누설분")
    else:
        put(r, 4, "미측정")
        put(r, 8, "라벨 '빈소켓_baseline'로 측정 필요")
    for c in (3, 5, 6, 7):
        put(r, c, "-")
    r += 2

    # 비교 요약
    if "기존" in group_avgs and "신규" in group_avgs:
        old_m, new_m = group_avgs["기존"], group_avgs["신규"]
        if baseline is not None:
            old_c, new_c = old_m - baseline, new_m - baseline
            delta = (new_c - old_c) / old_c * 100 if old_c else 0
            txt = (f"칩 단독 STOP 전류 (베이스라인 차감): 기존 {old_c:.3f} µA → "
                   f"신규 {new_c:.3f} µA  ({delta:+.1f}%)")
        else:
            delta = (new_m - old_m) / old_m * 100 if old_m else 0
            txt = f"측정 평균: 기존 {old_m:.3f} µA → 신규 {new_m:.3f} µA ({delta:+.1f}%)"
        c = ws.cell(row=r, column=2, value="비교: " + txt)
        c.font = Font(bold=True, size=11, color="1F3864")

    widths = {2: 16, 3: 6, 4: 15, 5: 11, 6: 11, 7: 19, 8: 30}
    from openpyxl.utils import get_column_letter
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    return group_avgs, baseline, OLD_LABELS, NEW_LABELS


FOOT_33 = ("* SCK = 오픈드레인 + MCU 내부 풀업 — 신호·전원 모두 3.3V로 레벨 "
           "일치 (입력단 관통 전류 없음, 외부 부품 0개). VDD 5.0V 조건은 "
           "'5V_수동' 시트 수동 기입 또는 레벨시프터 장착 시 '5V_자동' 시트")
FOOT_5 = ("* SCK/SDA = 레벨시프터 경유 (MCU 3.3V ↔ 칩 5V 변환) — GUI VDD "
          "조건 '5V (레벨시프터)' 선택, 펌웨어 `stop pd pp`(푸시풀 High) 사용")


def _label_key(lab: str):
    """시료 라벨 자연 정렬 (신규#2 < 신규#10)."""
    m = re.search(r"(\d+)\s*$", lab)
    return (re.sub(r"\d+\s*$", "", lab), int(m.group(1)) if m else -1)


def _build_simple_summary(wb, all_rows, rows33, rows5, tag: str):
    """심플 요약 (이름 런/직접 지정 파일용) — 칩별 데이터 표 + 판정만.

    2026-08-18 사용자 피드백: 새 캠페인 파일에 비교 캠페인의 풀 구조(수동
    시트·기존vs신규 매트릭스)가 따라오는 건 과함 — 데이터가 잘 보이는 표로."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("요약", 0)
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill("solid", fgColor="E2EFDA")
    bad_fill = PatternFill("solid", fgColor="FCE4E4")
    base_fill = PatternFill("solid", fgColor="EDEDED")
    CRIT_UA = 10.0

    ws.merge_cells("B2:G2")
    t = ws.cell(row=2, column=2, value=f"STOP 전류 결과 — {tag}")
    t.font = Font(bold=True, size=14, color="1F3864")
    ws.cell(row=3, column=2, value=(
        "판정 기준: VDD 5.0V STOP 누설 < 10 µA (칩 테스트 플랜). "
        "3.3V 행은 참고. 차감 = 측정 − 빈소켓 베이스라인.")).font = \
        Font(size=9, italic=True)

    def base_of(rows):
        return next((float(r["avg_uA"]) for k, r in rows.items()
                     if is_baseline(k)), None)

    base33, base5 = base_of(rows33), base_of(rows5)

    hdr = ["시료", "조건", "측정 평균 (µA)", "차감 (µA)", "판정", "측정 시각"]
    r0 = 5
    for i, h in enumerate(hdr):
        c = ws.cell(row=r0, column=2 + i, value=h)
        c.fill = head_fill
        c.font = head_font
        c.border = box
        c.alignment = Alignment(horizontal="center")

    # 5V 먼저, 각 조건 안에서 베이스라인 → 시료(자연 정렬)
    ordered = sorted(
        all_rows.items(),
        key=lambda kv: (kv[0][1] != "5.0", not is_baseline(kv[0][0]),
                        _label_key(kv[0][0])))
    r = r0 + 1
    sums: dict[str, list] = {}
    for (lab, vdd), d in ordered:
        avg = float(d["avg_uA"])
        base = base5 if vdd.startswith("5") else base33
        is_b = is_baseline(lab)
        sub = None if (is_b or base is None) else avg - base
        grp = _label_key(lab)[0].strip(" #_-") or "시료"   # 접두어별 그룹
        if is_b:
            verdict, fill = "—", base_fill
        elif vdd.startswith("5"):
            v = sub if sub is not None else avg
            verdict, fill = (("PASS", ok_fill) if v < CRIT_UA
                             else (f"FAIL ({v / CRIT_UA:.0f}배)", bad_fill))
            sums.setdefault((vdd, grp), []).append(v)
        else:
            verdict, fill = "참고", None
            if sub is not None:
                sums.setdefault((vdd, grp), []).append(sub)
        vals = [lab, f"{float(vdd):.1f} V", avg,
                sub if sub is not None else "—", verdict,
                str(d.get("datetime", ""))[:16]]
        for i, v in enumerate(vals):
            cell = ws.cell(row=r, column=2 + i, value=v)
            cell.border = box
            cell.alignment = Alignment(horizontal="center")
            if isinstance(v, float):
                cell.number_format = "0.000"
            if is_b:
                cell.fill = base_fill
            elif i == 4 and fill:
                cell.fill = fill
                cell.font = Font(bold=True)
        r += 1

    r += 1
    for (vdd, grp) in sorted(sums, key=lambda k: (k[0] != "5.0", k[1])):
        vals = sums[(vdd, grp)]
        n_pass = sum(1 for v in vals if v < CRIT_UA)
        line = (f"{float(vdd):.1f} V · {grp}: {len(vals)}개, 차감 평균 "
                f"{sum(vals) / len(vals):.3f} µA")
        if vdd.startswith("5"):
            line += f", PASS {n_pass}/{len(vals)}"
        ws.cell(row=r, column=2, value=line).font = Font(bold=True, size=10)
        r += 1

    widths = {2: 16, 3: 9, 4: 15, 5: 12, 6: 16, 7: 17}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def generate(run: str | None = None, csv: str | None = None,
             out: str | None = None):
    """CSV → 보고표 생성. GUI가 측정 직후 자동 호출 (수동: main).

    run = 테스트 그룹(런) 이름 — 그룹별로 stop_current_<run>.csv /
    STOP_current_report_<run>.xlsx 로 완전 분리 (None = 기본 캠페인).
    csv/out = **명시 경로 지정** (GUI '보고표 파일' 선택, 2026-08-18) —
    지정 시 run 이름 규칙 대신 그 파일을 사용.
    CSV의 vdd 컬럼(없으면 3.3 취급)으로 조건을 분리 — 3.3V 행은 '3V3_자동',
    5V 행(레벨시프터 자동화)은 '5V_자동' 시트로. 5V 행이 없으면 5V_자동
    시트는 생성하지 않는다. 수동 시트의 시료 행은 CSV·기존 입력에 등장한
    라벨로 동적 구성 ('기존*'=기존 그룹, 그 외=신규 그룹)."""
    cpath = Path(csv) if csv else csv_path(run)
    opath = Path(out) if out else out_path(run)
    all_rows = load_rows(cpath)
    rows33, rows5 = {}, {}
    for (lab, v), r in all_rows.items():
        (rows5 if v.startswith("5") else rows33)[lab] = r

    # ---- 심플 모드 (이름 런/직접 지정 파일): 데이터 표 + 조건별 시트만 ----
    # 기본 캠페인(무인자)만 제출 확정 풀 구조(수동 시트·비교 요약) 유지.
    if run or csv or out:
        from openpyxl import Workbook
        tag = run or Path(out or cpath).stem
        wb = Workbook()
        built = False
        if rows33:
            _build_auto_sheet(
                wb, rows33, sheet_name="3V3_자동",
                title="STOP Procedure — VDD 3.3V 자동 측정",
                sck_val="0/3.3V*", vdd_val="3.3 V", foot=FOOT_33, first=True)
            built = True
        if rows5:
            _build_auto_sheet(
                wb, rows5, sheet_name="5V_자동",
                title="STOP Procedure — VDD 5.0V 자동 측정",
                sck_val="0/5V*", vdd_val="5.0 V", foot=FOOT_5,
                first=not built)
            built = True
        if not built:
            wb.active.title = "데이터 없음"
        _build_simple_summary(wb, all_rows, rows33, rows5, tag)
        try:
            wb.save(opath)
        except PermissionError:
            sys.exit(f"저장 실패 — {opath.name}이 엑셀에 열려 있음. 닫고 재실행.")
        print(f"보고표 갱신(심플): {opath}")
        print(f"  행 {len(all_rows)}개 (3.3V {len(rows33)} / 5V {len(rows5)})")
        return

    from openpyxl import Workbook

    # 기존 보고표의 수동 시트 입력값 보존 (재생성 시 사용자가 손으로
    # 넣은 판독값이 날아가지 않게)
    manual_saved = _load_manual_values(opath)

    # 수동 시트 시료 행 = CSV 라벨 ∪ 기존 수동 입력 라벨 (없으면 기본 틀)
    seen = sorted({lab for (lab, _v) in all_rows})
    old_labels = [l for l in seen if l.startswith("기존")]
    new_labels = [l for l in seen
                  if not l.startswith("기존") and not is_baseline(l)]
    for (_sheet, _sec, lab) in manual_saved:
        if is_baseline(lab):
            continue
        tgt = old_labels if lab.startswith("기존") else new_labels
        if lab not in tgt:
            tgt.append(lab)
    if not old_labels and not new_labels:
        old_labels = list(MANUAL_ROWS[:3])
        new_labels = list(MANUAL_ROWS[3:6])
    old_labels.sort()
    new_labels.sort()

    wb = Workbook()
    group_avgs, baseline, OLD_LABELS, NEW_LABELS = _build_auto_sheet(
        wb, rows33, sheet_name="3V3_자동",
        title="STOP Procedure — VDD 3.3V 자동 측정 (GUI 연동, 패키지 시료 비교)",
        sck_val="0/3.3V*", vdd_val="3.3 V", foot=FOOT_33, first=True)
    if rows5:
        _build_auto_sheet(
            wb, rows5, sheet_name="5V_자동",
            title="STOP Procedure — VDD 5.0V 자동 측정 (레벨시프터 경유)",
            sck_val="0/5V*", vdd_val="5.0 V", foot=FOOT_5)

    for name, cfg in MANUAL_SHEETS.items():
        _build_manual_sheet(wb, name, cfg, manual_saved,
                            old_labels, new_labels)

    # 요약 시트 (맨 앞): 3.3V 자동 그룹평균(차감) + 5V 수동 STOP 인용
    base5 = _manual_mean(manual_saved, "5V_수동", "stop", ["빈소켓_baseline"])
    new5 = _manual_mean(manual_saved, "5V_수동", "stop", new_labels)
    old5 = _manual_mean(manual_saved, "5V_수동", "stop", old_labels)

    # 5V 자동측정(캡+풀업/레벨시프터) 폴백 — 수동 시트 입력이 없으면 CSV의
    # 5V 행에서 그룹 평균 산출 (요약 시트 인용용)
    def _auto_mean(rows, labels):
        vals = [float(rows[lb]["avg_uA"]) for lb in labels if lb in rows]
        return sum(vals) / len(vals) if vals else None

    base5a = next((float(r["avg_uA"]) for k, r in rows5.items()
                   if is_baseline(k)), None)
    if new5 is None:
        n5a = _auto_mean(rows5, new_labels)
        new5 = (n5a - (base5a or 0)) if n5a is not None else None
    else:
        new5 -= (base5 or 0)
    if old5 is None:
        o5a = _auto_mean(rows5, old_labels)
        old5 = (o5a - (base5a or 0)) if o5a is not None else None
    else:
        old5 -= (base5 or 0)
    sumdata = {
        "old_33": (group_avgs["기존"] - baseline)
        if ("기존" in group_avgs and baseline is not None) else None,
        "new_33": (group_avgs["신규"] - baseline)
        if ("신규" in group_avgs and baseline is not None) else None,
        "new_5": new5,
        "old_5": old5,
        "new_5n": _manual_mean(manual_saved, "5V_수동", "normal", new_labels),
        "old_5n": _manual_mean(manual_saved, "5V_수동", "normal", old_labels),
        "new_33n": _manual_mean(manual_saved, "3V3_수동", "normal", new_labels),
        "old_33n": _manual_mean(manual_saved, "3V3_수동", "normal", old_labels),
    }
    if run:
        _build_summary_generic(wb, sumdata, run)
    else:
        _build_summary_sheet(wb, sumdata)   # 기본 캠페인 = 제출 확정본 유지

    try:
        wb.save(opath)
    except PermissionError:
        sys.exit(f"저장 실패 — {opath.name}이 엑셀에 열려 있음. 닫고 재실행.")
    print(f"보고표 갱신: {opath}")
    print(f"  3.3V: 기존 {len(OLD_LABELS)}개 + 신규 {len(NEW_LABELS)}개, 베이스라인 "
          f"{'있음' if baseline is not None else '없음(빈소켓_baseline 필요)'}"
          + (f" / 5V 자동: {len(rows5)}행" if rows5 else ""))


def _build_summary_generic(wb, data: dict, run: str):
    """이름 지정 런(새 테스트 그룹)용 요약 — 판정은 기준값으로 자동 산출.
    (기본 캠페인의 요약은 제출 확정 문구가 있어 _build_summary_sheet 유지)"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("요약", 0)
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill("solid", fgColor="E2EFDA")
    bad_fill = PatternFill("solid", fgColor="FCE4E4")
    CRIT_UA = 10.0

    ws.merge_cells("B2:G2")
    t = ws.cell(row=2, column=2, value=f"STOP 전류 시험 요약 — {run}")
    t.font = Font(bold=True, size=14, color="1F3864")
    ws.cell(row=4, column=2, value=(
        "판정 기준: 칩 테스트 플랜 'STOP' — Shutdown 모드 누설 VDD < 10 µA "
        "(VDD 5.0V, SCK High 유지)")).font = Font(bold=True, size=10)

    hdr = ["그룹", "5V STOP (µA, 차감)", "판정", "3.3V STOP (µA, 차감)",
           "5V Normal (mA)", "3.3V Normal (mA)"]
    r0 = 6
    for i, h in enumerate(hdr):
        c = ws.cell(row=r0, column=2 + i, value=h)
        c.fill = head_fill
        c.font = head_font
        c.border = box
        c.alignment = Alignment(horizontal="center")

    def fmt(v):
        return f"{v:.3f}" if isinstance(v, float) else "측정 대기"

    def verdict(v):
        if not isinstance(v, float):
            return ("측정 대기", None)
        return ("PASS", ok_fill) if v < CRIT_UA else \
            (f"FAIL (기준 {v / CRIT_UA:.0f}배)", bad_fill)

    rows = [("신규 그룹", data.get("new_5"), data.get("new_33"),
             data.get("new_5n"), data.get("new_33n")),
            ("기존 그룹", data.get("old_5"), data.get("old_33"),
             data.get("old_5n"), data.get("old_33n"))]
    for j, (gname, v5, v33, n5, n33) in enumerate(rows):
        rr = r0 + 1 + j
        vtxt, vfill = verdict(v5)
        vals = [gname, fmt(v5), (vtxt, vfill), fmt(v33), fmt(n5), fmt(n33)]
        for i, val in enumerate(vals):
            cell = ws.cell(row=rr, column=2 + i)
            if isinstance(val, tuple):
                cell.value = val[0]
                if val[1]:
                    cell.fill = val[1]
                cell.font = Font(bold=True)
            else:
                cell.value = val
            cell.border = box
            cell.alignment = Alignment(horizontal="center")

    notes = [
        "단위: STOP 표 = µA (DMM이 mA 표시면 ×1000 환산 기입) / Normal 표 = mA.",
        "그룹 규칙: 라벨 '기존*' = 기존 그룹, 빈소켓/baseline = 베이스라인, 그 외 = 신규 그룹.",
        "근거 데이터: '3V3_자동'(GUI 자동측정) · '3V3_수동' · '5V_수동' 시트 참조.",
    ]
    for i, s in enumerate(notes):
        ws.cell(row=r0 + 4 + i, column=2, value=s).font = Font(size=9,
                                                               italic=True)
    widths = {2: 14, 3: 20, 4: 18, 5: 20, 6: 16, 7: 16}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def main():
    arg = sys.argv[1].strip() if len(sys.argv) > 1 else None
    if arg and arg.lower().endswith(".xlsx"):
        out = Path(arg)                       # 파일 직접 지정 모드
        generate(csv=str(out.with_suffix(".csv")), out=str(out))
    else:
        generate(arg or None)


if __name__ == "__main__":
    main()
