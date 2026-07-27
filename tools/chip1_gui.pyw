"""
CHIP1 ADC test — GUI front-end (v4 펌웨어 이식판).

tools/chip1_autotest.py(신판)를 import해서 시리얼/엑셀 로직을 재사용한다.
구 GUI(레거시 2보드 셋업용)와 같은 구성이지만:
  - 보드 1개 (Nucleo 통합) — 보드 식별/칩선택 없음, id(0x9210) 확인으로 대체
  - DAC 채널 매핑 반전: ch1(PA4)=AINP, ch2(PA5)=AINN  ← 구셋업과 반대!
  - 내장 DAC cal 계수 자동 입력 (부팅마다 필요, autotest의 DAC_CAL 사용)
  - 기본 엑셀 = CHIP1_ADC_validation.xlsx (없으면 template/report_template.xlsx
    복사로 자동 생성, 템플릿 원본은 수정 안 함)

더블클릭 실행 (pythonw, 콘솔 없음 — 출력은 아래 로그 창에).
"""

import builtins
import csv
import math
import os
import queue
import re
import statistics
import sys
import threading
import time
import traceback
from copy import copy
from datetime import datetime
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
from tkinter.scrolledtext import ScrolledText

# 어디서 실행돼도 tools/의 신판 autotest를 import
sys.path.insert(0, str(Path(__file__).resolve().parent))
import chip1_autotest as g
from chamber import Chamber, MockChamber, PROFILES, ChamberError
from temp_sweep import SweepConfig, TempSweep

SWEEP_TEMPS_DEFAULT = "-40,-20,0,10,25,40,70,85"
SWEEP_CSV_FIELDS = [                       # block diagram p.5 CSV 포맷 (2안)
    "timestamp", "set_temp_C", "chamber_temp_C", "mode",
    "VDD_V", "VREF_V", "vref_valid",       # vref_valid: SWEEP_VREF_RANGE 내 여부
    "reg_settings",                        # 마지막 wr 명령 최대 3개 (;구분)
    "rd_mean", "rd_stdev", "rd_count", "dut_block", "sample_csv",
]
# VREF(=VDDA, 칩 내부 LDO 3.0V) 정상 범위 — 밖이면 칩 미장착/미개조 보드/
# 접촉 불량 의심 (SPEC.md §8 "VDDA/REFIN 노드 확정")
SWEEP_VREF_RANGE = (2.9, 3.1)

DUT_RE = re.compile(r"DUT\s*#\s*(\d+)")

PGA_GAIN = 64  # report row 8 'PGA (V/V)'; 0x03 = 0x02 고정 전제


# ---------------- result summary (엑셀 리포트 수식과 동일) ----------------

def summarize_internal(values: list[int]):
    """Internal Short: (std, enob). STD = STDEV.S; ENOB = LOG2(0.9*2^24/STD)."""
    if len(values) < 2:
        return None, None
    std = statistics.stdev(values)
    enob = math.log2(0.9 * 2 ** 24 / std) if std > 0 else None
    return std, enob


def summarize_channel_a(values: list[int], ainp_uv: int, ainn_uv: int):
    """Channel A: (avg, vin_calc, vin_actual, accuracy).
    Vin(calc) = 3/2^23 * AVG / PGA;  Vin(actual) = (AINP-AINN) µV.
    ⚠ 신판: AINP=ch1, AINN=ch2 (구 GUI는 dac2-dac1이었음 — 매핑 반전)."""
    if not values:
        return None, None, None, None
    avg = statistics.fmean(values)
    vin_calc = 3 / 2 ** 23 * avg / PGA_GAIN
    vin_actual = (ainp_uv - ainn_uv) / 1e6
    accuracy = vin_calc / vin_actual if vin_actual else None
    return avg, vin_calc, vin_actual, accuracy


MODES = [
    ("internal",  "Internal Short (0x04 = 0x60)", "0x60"),
    ("channel_a", "Channel A (0x04 = 0x00)",      "0x00"),
]
MODE_LABELS = {key: label for key, label, _ in MODES}

# -------- 2점 캘리브레이션 수식 --------
#
# 보정 0 상태에서 지시 s → 실측 m 이 선형이라 가정: m = A*s + B
#   두 점 (s1,m1), (s2,m2)에서  A = (m2-m1)/(s2-s1),  B = m1 - A*s1
# 목표는 지시 s가 실측 s로 나오게 하는 corrected c(s) = (s - B)/A.
# 펌웨어 형태 c(s) = (s + offset)*(1 + ppm/1e6) 와 계수 비교하면:
#   1 + ppm/1e6 = 1/A   →  gain_ppm  = round((1/A - 1) * 1e6)
#   offset*(1/A) = -B/A →  offset_uV = round(-B)
CAL_VERIFY_UV = 1750000    # 검증 중간점
CAL_TOL_UV = 2000          # 검증 목표 ±2mV
CAL_LOW_DEFAULT = 500000   # 저점 기본 0.5V
CAL_HIGH_DEFAULT = 3000000  # 고점 기본 3.0V (12bit 범위 안에서 넓을수록 기울기 정확)


def cal_two_point(s1_uv: int, m1_uv: float, s2_uv: int, m2_uv: float):
    """2점 (지시, 실측)µV → (offset_uV, gain_ppm) 정수 계수."""
    a = (m2_uv - m1_uv) / (s2_uv - s1_uv)
    if not 0.5 < a < 2.0:      # 기울기가 상식 밖이면 측정/배선 오류
        raise ValueError(f"기울기 A={a:.4f} — 실측값이 이상해요 (배선/채널 확인)")
    b = m1_uv - a * s1_uv
    gain_ppm = round((1.0 / a - 1.0) * 1e6)
    offset_uv = round(-b)
    return offset_uv, gain_ppm


# ---------------- Excel: DUT 블록 자동 확장 (구 GUI에서 이식) ----------------

def _scan_duts(ws) -> dict[int, int]:
    duts = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=g.HEADER_ROW, column=c).value
        if isinstance(v, str):
            m = DUT_RE.search(v)
            if m:
                duts[int(m.group(1))] = c
    return duts


def _copy_block(ws, src_col: int, dst_col: int, new_dut: int):
    """DUT 블록(2열) 템플릿 복제: 헤더/모드행/수식/스타일/병합. 데이터는 비움."""
    from openpyxl.cell.cell import MergedCell
    from openpyxl.formula.translate import Translator
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.cell_range import CellRange

    shift = dst_col - src_col
    for rng in list(ws.merged_cells.ranges):
        if (rng.min_col >= src_col and rng.max_col <= src_col + 1
                and rng.min_row >= g.HEADER_ROW
                and rng.max_row < g.DATA_START_ROW):
            new_rng = CellRange(min_col=rng.min_col + shift, min_row=rng.min_row,
                                max_col=rng.max_col + shift, max_row=rng.max_row)
            if not any(new_rng.issubset(ex) or ex.issubset(new_rng)
                       for ex in ws.merged_cells.ranges):
                ws.merge_cells(new_rng.coord)

    last_data_row = g.DATA_START_ROW + g.N_SAMPLES - 1
    for off in (0, 1):
        sc, dc = src_col + off, dst_col + off
        src_dim = ws.column_dimensions[get_column_letter(sc)]
        ws.column_dimensions[get_column_letter(dc)].width = src_dim.width
        for r in range(g.HEADER_ROW, last_data_row + 1):
            s = ws.cell(row=r, column=sc)
            d = ws.cell(row=r, column=dc)
            d._style = copy(s._style)
            if r >= g.DATA_START_ROW or isinstance(d, MergedCell):
                continue
            v = s.value
            if isinstance(v, str) and v.startswith("="):
                d.value = Translator(v, origin=s.coordinate).translate_formula(
                    d.coordinate)
            elif r == g.HEADER_ROW and isinstance(v, str):
                d.value = DUT_RE.sub(f"DUT#{new_dut}", v)
            elif not isinstance(s, MergedCell):
                d.value = v


def norm_port(s: str) -> str:
    """포트 입력 정규화: '8' -> 'COM8' (숫자만 치는 실수 흡수)."""
    s = (s or "").strip()
    return f"COM{s}" if s.isdigit() else s


def create_sweep_sheet(excel: Path, label: str = "") -> str:
    """스윕 전용 시트 생성: 기존 레이아웃 시트를 복제해 데이터만 비운 새 시트.

    상온 검증 시트(원본)는 건드리지 않고, 스윕마다 'SW_MMDD_HHMM[_라벨]'
    시트가 하나씩 생긴다 (시트 안에서는 블록 = 온도). 반환: 시트명.
    """
    from openpyxl import load_workbook

    while True:
        try:
            wb = load_workbook(excel)
            break
        except PermissionError:
            input("엑셀 파일이 열려 있어요! Excel을 닫은 다음 OK를 눌러 주세요.")
    template = wb[g.SHEET_NAME] if g.SHEET_NAME else wb.active
    ws = wb.copy_worksheet(template)

    name = datetime.now().strftime("SW_%m%d_%H%M")
    if label:
        name += "_" + re.sub(r"[\[\]:*?/\\'\s]+", "_", label.strip())
    base = name[:28]                      # 엑셀 시트명 31자 제한 + 중복 여유
    name, i = base, 1
    existing = [s for s in wb.sheetnames if s != ws.title]
    while name in existing:
        name = f"{base}_{i}"
        i += 1
    ws.title = name

    # 복제된 데이터/온도라벨 제거 -> 빈 블록 템플릿 상태로
    for _dut, c in sorted(_scan_duts(ws).items(), key=lambda kv: kv[1]):
        for col in (c, c + 1):
            for r in range(g.DATA_START_ROW, g.DATA_START_ROW + g.N_SAMPLES):
                if ws.cell(row=r, column=col).value is not None:
                    ws.cell(row=r, column=col).value = None
        hcell = ws.cell(row=g.HEADER_ROW, column=c)
        if isinstance(hcell.value, str):
            hcell.value = re.sub(r"\s*@.*$", "", hcell.value)

    while True:
        try:
            wb.save(excel)
            return name
        except PermissionError:
            input("엑셀 파일이 열려 있어요! Excel을 닫은 다음 OK를 눌러 주세요.")


class QueueWriter:
    def __init__(self, q: queue.Queue, mirror: list | None = None):
        self.q = q
        self.mirror = mirror

    def write(self, s: str):
        if s:
            self.q.put(s)
            if self.mirror is not None:
                self.mirror.append(s)

    def flush(self):
        pass


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("CHIP1 ADC Test (v4 board)")
        root.minsize(560, 500)
        self.q: queue.Queue[str] = queue.Queue()

        frm = ttk.Frame(root, padding=10)
        frm.grid(sticky="nsew")
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)

        self._row = 0

        notice = ttk.LabelFrame(frm, text="준비사항 (read me first!)", padding=6)
        notice.grid(row=self._row, column=0, columnspan=3, sticky="we", pady=(0, 8))
        ttk.Label(notice, justify="left", text=(
            "1) 테스트/스윕 중 결과 엑셀(results\\CHIP1_ADC_validation.xlsx)을\n"
            "    열어두지 마세요 — 기록 시점에 잠겨 있으면 팝업으로 멈춥니다 (무인 스윕 주의)\n"
            "2) Tera Term 등 터미널은 닫고 실행하세요 — 보드 COM과 챔버 COM 모두\n"
            "    한 프로그램만 쓸 수 있습니다 (수동 확인할 때만 열기)\n"
            "3) 캘 계수(DAC/meas)는 config\\에서 자동 적용 — Nucleo 보드나 배선이\n"
            "    바뀌면 아래 DAC/ADC Calibration 섹션에서 재캘 필요\n"
            "4) VDD/VREF 측정은 개조 표시가 있는 센서보드 전용입니다.\n"
            "    미개조 보드는 VREF가 1.2V대로 표시 — 고장이 아니라 측정 불가 상태\n"
            "    (사용하려면 C1-REFIN 쇼트 개조 필요, docs 참조)\n"
            "5) 스윕 결과 = 새 엑셀 시트(SW_...) + results\\sweeps\\<런폴더>\\\n"
            "6) 파이썬 라이브러리: pip install -r requirements.txt (tools\\_vendor 있으면 자동 사용)"
        )).grid(sticky="w")
        self._row += 1

        def add_entry(label: str, default: str) -> tk.StringVar:
            ttk.Label(frm, text=label).grid(row=self._row, column=0, sticky="w")
            var = tk.StringVar(value=default)
            ttk.Entry(frm, textvariable=var, width=20).grid(
                row=self._row, column=1, sticky="we", pady=2)
            self._row += 1
            return var

        # ⚠ 채널 매핑: ch1(PA4)=AINP, ch2(PA5)=AINN — 구 GUI와 반대!
        self.ainp = add_entry("dac set 1 = AINP µV  (PA4)", str(g.AINP_UV))
        self.ainn = add_entry("dac set 2 = AINN µV  (PA5)", str(g.AINN_UV))
        self.nsamp = add_entry("How many samples", str(g.N_SAMPLES))
        self.settle = add_entry("Settle wait sec (드리프트 시 60)", "0")
        self.dut = add_entry("DUT #  (blank = next empty in Excel)", "")
        self.port = add_entry("COM port  (blank = auto)", "")

        ttk.Label(frm, text="Excel report").grid(row=self._row, column=0, sticky="w")
        self.excel = tk.StringVar(value=str(g.EXCEL_PATH))
        self.excel_box = ttk.Combobox(frm, textvariable=self.excel, width=24)
        self.excel_box.grid(row=self._row, column=1, sticky="we", pady=2)

        def refresh_xlsx_list():
            names = [str(g.EXCEL_PATH)]
            names += sorted(str(p) for p in g.PROJECT_ROOT.glob("*.xlsx")
                            if not p.name.startswith("~$") and str(p) not in names)
            self.excel_box["values"] = names

        self.excel_box.configure(postcommand=refresh_xlsx_list)
        refresh_xlsx_list()

        def browse_xlsx():
            path = filedialog.askopenfilename(
                title="엑셀 리포트 선택", initialdir=g.PROJECT_ROOT,
                filetypes=[("Excel 파일", "*.xlsx")])
            if path:
                self.excel.set(path)

        ttk.Button(frm, text="찾아보기…", command=browse_xlsx, width=10).grid(
            row=self._row, column=2, padx=(4, 0))
        self._row += 1

        self.mode_vars: dict[str, tk.BooleanVar] = {}
        for key, label, _ in MODES:
            var = tk.BooleanVar(value=True)
            ttk.Checkbutton(frm, text=label, variable=var).grid(
                row=self._row, column=0, columnspan=2, sticky="w")
            self.mode_vars[key] = var
            self._row += 1

        # -------- 오른쪽 열 컨테이너 (캘 2종 + 스윕 + 로그) --------
        self._right = ttk.Frame(frm)

        # -------- DAC Calibration (2점 위저드) — 오른쪽 --------
        calf = ttk.LabelFrame(self._right, text="DAC Calibration (2-point wizard)",
                              padding=6)
        calf.pack(side="top", fill="x")
        ttk.Label(calf, text="Channel").grid(row=0, column=0, sticky="w")
        self.cal_ch = tk.StringVar(value="1")
        ttk.Combobox(calf, textvariable=self.cal_ch, values=("1", "2"),
                     width=4, state="readonly").grid(row=0, column=1, padx=(2, 10))
        ttk.Label(calf, text="저점 µV").grid(row=0, column=2, sticky="w")
        self.cal_low = tk.StringVar(value=str(CAL_LOW_DEFAULT))
        ttk.Entry(calf, textvariable=self.cal_low, width=9).grid(row=0, column=3, padx=(2, 10))
        ttk.Label(calf, text="고점 µV").grid(row=0, column=4, sticky="w")
        self.cal_high = tk.StringVar(value=str(CAL_HIGH_DEFAULT))
        ttk.Entry(calf, textvariable=self.cal_high, width=9).grid(row=0, column=5, padx=(2, 10))
        self.cal_btn = ttk.Button(calf, text="Calibrate…", command=self.cal_clicked)
        self.cal_btn.grid(row=0, column=6, padx=(4, 0))
        ttk.Label(calf, foreground="gray", text=(
            "멀티미터를 해당 채널 출력(ch1=D1, ch2=D13)과 GND에 연결하고 시작하세요. "
            f"완료 시 보드 플래시+json에 저장됩니다.")).grid(
            row=1, column=0, columnspan=8, sticky="w", pady=(4, 0))

        # -------- ADC Calibration (meas) — 오른쪽 --------
        adcf = ttk.LabelFrame(self._right, text="ADC Calibration (meas)", padding=6)
        adcf.pack(side="top", fill="x", pady=(6, 0))
        ttk.Label(adcf, text="Channel").grid(row=0, column=0, sticky="w")
        self.acal_ch = tk.StringVar(value="VREF (A4)")
        ttk.Combobox(adcf, textvariable=self.acal_ch, state="readonly", width=10,
                     values=("VREF (A4)", "VDD (A3)")).grid(row=0, column=1, padx=(2, 10))
        ttk.Label(adcf, text="저점 µV").grid(row=0, column=2, sticky="w")
        self.acal_low = tk.StringVar(value=str(CAL_LOW_DEFAULT))
        ttk.Entry(adcf, textvariable=self.acal_low, width=9).grid(row=0, column=3, padx=(2, 10))
        ttk.Label(adcf, text="고점 µV").grid(row=0, column=4, sticky="w")
        self.acal_high = tk.StringVar(value=str(CAL_HIGH_DEFAULT))
        ttk.Entry(adcf, textvariable=self.acal_high, width=9).grid(row=0, column=5, padx=(2, 10))
        self.selfcal_btn = ttk.Button(adcf, text="Self-cal (배선 불필요)",
                                      command=self.self_cal_clicked)
        self.selfcal_btn.grid(row=0, column=6, padx=(4, 0))
        self.acal_btn = ttk.Button(adcf, text="Advanced…", width=10,
                                   command=self.adc_cal_clicked)
        self.acal_btn.grid(row=0, column=7, padx=(4, 0))
        self.acal_status = tk.StringVar(value=self._meas_cal_summary())
        ttk.Label(adcf, textvariable=self.acal_status, foreground="gray").grid(
            row=1, column=0, columnspan=8, sticky="w", pady=(4, 0))

        # -------- Temperature Sweep (chamber) — 오른쪽 --------
        swf = ttk.LabelFrame(self._right, text="Temperature Sweep (chamber)",
                             padding=6)
        swf.pack(side="top", fill="x", pady=(6, 0))
        ttk.Label(swf, text="온도 리스트(°C)").grid(row=0, column=0, sticky="w")
        self.sw_temps = tk.StringVar(value=SWEEP_TEMPS_DEFAULT)
        ttk.Entry(swf, textvariable=self.sw_temps, width=34).grid(
            row=0, column=1, columnspan=3, sticky="we", padx=(2, 6))
        ttk.Label(swf, text="포화(분)").grid(row=1, column=0, sticky="w")
        self.sw_soak = tk.StringVar(value="10")
        ttk.Entry(swf, textvariable=self.sw_soak, width=6).grid(row=1, column=1, sticky="w", padx=(2, 6))
        ttk.Label(swf, text="허용오차(±°C)").grid(row=1, column=2, sticky="e")
        self.sw_tol = tk.StringVar(value="1.0")
        ttk.Entry(swf, textvariable=self.sw_tol, width=6).grid(row=1, column=3, sticky="w", padx=(2, 6))
        ttk.Label(swf, text="챔버 프로파일").grid(row=2, column=0, sticky="w")
        self.sw_profile = tk.StringVar(value="SH662_RS485")
        ttk.Combobox(swf, textvariable=self.sw_profile, state="readonly", width=13,
                     values=list(PROFILES) + ["MOCK"]).grid(row=2, column=1, sticky="w", padx=(2, 6))
        ttk.Label(swf, text="챔버 포트").grid(row=2, column=2, sticky="e")
        self.sw_port = tk.StringVar(value="")
        ttk.Entry(swf, textvariable=self.sw_port, width=8).grid(row=2, column=3, sticky="w", padx=(2, 6))
        ttk.Label(swf, text="라벨(EVM/칩, 선택)").grid(row=3, column=0, sticky="w")
        self.sw_label = tk.StringVar(value="")
        ttk.Entry(swf, textvariable=self.sw_label, width=14).grid(
            row=3, column=1, sticky="w", padx=(2, 6))
        ttk.Label(swf, text="종료 후").grid(row=3, column=2, sticky="e")
        self.sw_end = tk.StringVar(value="25C 복귀 후 정지")
        ttk.Combobox(swf, textvariable=self.sw_end, state="readonly", width=14,
                     values=("25C 복귀 후 정지", "즉시 정지", "유지(수동)")).grid(
            row=3, column=3, columnspan=2, sticky="w", padx=(2, 0))
        ttk.Label(swf, foreground="gray",
                  text="스윕마다 새 엑셀 시트(SW_날짜_시간_라벨) 생성 — 상온 검증 시트는 불변").grid(
            row=4, column=0, columnspan=5, sticky="w")
        self.sweep_btn = ttk.Button(swf, text="Start sweep", command=self.sweep_clicked)
        self.sweep_btn.grid(row=0, column=4, rowspan=2, sticky="nswe", padx=(4, 0))
        self.sweep_stop_btn = ttk.Button(swf, text="중단…", state="disabled",
                                         command=self.sweep_stop_clicked)
        self.sweep_stop_btn.grid(row=2, column=4, sticky="we", padx=(4, 0))
        self.sw_selfcal = tk.BooleanVar(value=True)
        ttk.Checkbutton(swf, text="온도별 자동 셀프캘 (meas cal — ADC 온도 드리프트 대응)",
                        variable=self.sw_selfcal).grid(
            row=5, column=0, columnspan=5, sticky="w")
        self.sw_status = tk.StringVar(value="대기 중")
        ttk.Label(swf, textvariable=self.sw_status, foreground="dark blue",
                  justify="left").grid(row=6, column=0, columnspan=5, sticky="w", pady=(4, 0))
        # (스윕 섹션은 오른쪽 열이라 왼쪽 self._row를 소비하지 않음)

        # 보드 상태 표시줄 (연결 감지 — UID + 플래시 캘 유무)
        self.board_status = tk.StringVar(value="보드 상태: 미확인")
        ttk.Label(frm, textvariable=self.board_status, foreground="dark green").grid(
            row=self._row, column=0, columnspan=2, sticky="w")
        ttk.Button(frm, text="보드 확인", width=10, command=self.check_board_clicked).grid(
            row=self._row, column=2, sticky="e")
        self._row += 1

        self.run_btn = ttk.Button(frm, text="Run test", command=self.run_clicked)
        self.run_btn.grid(row=self._row, column=0, columnspan=2, pady=6, sticky="we")
        self._row += 1
        self.run_stop_btn = ttk.Button(frm, text="Run 중단", state="disabled",
                                       command=self.run_stop_clicked)
        self.run_stop_btn.grid(row=self._row, column=0, columnspan=2, sticky="we")
        self._row += 1
        self._stop_run = False        # Run 중단 요청 플래그
        self._active_ser = None       # 진행 중인 시리얼 (ESC 주입용)

        # 오른쪽 열 = 스윕 섹션(위) + 로그(아래) — 왼쪽 열 0~2, 오른쪽 열 3
        self.log_box = ScrolledText(self._right, width=62, height=18,
                                    state="disabled", font=("Consolas", 9))
        self.log_box.pack(side="top", fill="both", expand=True, pady=(8, 0))
        self._right.grid(row=0, column=3, rowspan=self._row, sticky="nsew",
                         padx=(10, 0))
        frm.rowconfigure(self._row - 1, weight=1)
        frm.columnconfigure(1, weight=0)
        frm.columnconfigure(3, weight=1)
        root.minsize(1180, 640)

        root.after(100, self._poll_log)
        root.after(400, self.check_board_clicked)   # 기동 시 자동 1회 감지

    # ---------------- board identity / detection ----------------

    def check_board_clicked(self):
        threading.Thread(target=self._check_board_worker, daemon=True).start()

    def _check_board_worker(self):
        """연결된 보드의 UID + 플래시 캘 유무 확인 → 상태줄 갱신 + 미캘 안내.
        진실은 보드 플래시. 캘 있으면 UID를 이 PC 레지스트리에 기록."""
        try:
            port = norm_port(self.port.get()) or g.find_port()
            with g.open_port(port) as ser:
                st = g.board_status(ser)
        except (g.serial.SerialException, SystemExit, g.CliError):
            self.root.after(0, lambda: self.board_status.set(
                "보드 상태: 미연결 (보드 확인 버튼으로 재시도)"))
            return

        uid_short = (st["uid"][-8:] if st["uid"] else "구펌웨어(uid 없음)")
        fch = st["flash_ch"]

        def cal_txt(ch):
            off, ppm = st["cal"][ch]
            return f"ch{ch}: {off:+d}µV/{ppm:+d}ppm"

        if st["flash_cal"]:
            g.registry_note(st["uid"], flash_cal=True,
                            cal_ch1=list(st["cal"][1]), cal_ch2=list(st["cal"][2]))
            txt = (f"보드 상태: 캘 OK ✓ (양 채널)  UID …{uid_short}\n"
                   f"  적용 계수 — {cal_txt(1)}, {cal_txt(2)} (보드 플래시)")
            self.root.after(0, lambda: self.board_status.set(txt))
            return

        missing = [ch for ch in (1, 2) if not fch[ch]]
        saved = [ch for ch in (1, 2) if fch[ch]]
        miss_txt = "/".join(f"ch{c}" for c in missing)
        detail = ("  저장된 계수 — " + ", ".join(cal_txt(c) for c in saved) + "\n"
                  if saved else "")
        txt = f"보드 상태: ⚠ {miss_txt} 미캘  UID …{uid_short}\n{detail}".rstrip()
        self.root.after(0, lambda: self.board_status.set(txt))
        # 미캘 채널을 콕 집어 안내 + 위저드 채널을 그 채널로 미리 선택
        first = missing[0]
        self.root.after(0, lambda: self.cal_ch.set(str(first)))
        self.root.after(0, lambda: messagebox.showwarning(
            "보드 캘리브레이션 필요",
            f"이 보드는 {miss_txt} 캘리브레이션이 안 돼 있습니다.\n\n"
            f"DAC Calibration 섹션에서 Calibrate…를 실행하세요\n"
            f"(채널 {first}이 미리 선택돼 있습니다).\n\n"
            "완료되면 보드에 자동 저장되어, 이후 어느 PC에\n"
            "꽂아도 다시 할 필요가 없습니다."))

    # ---------------- log plumbing ----------------

    def _poll_log(self):
        try:
            while True:
                s = self.q.get_nowait()
                self.log_box.configure(state="normal")
                self.log_box.insert("end", s)
                self.log_box.see("end")
                self.log_box.configure(state="disabled")
        except queue.Empty:
            pass
        self.root.after(100, self._poll_log)

    def gui_ask_volts(self, prompt: str) -> float | None:
        """워커 스레드에서 실측 전압(V) 입력받기. 숫자 + 0~3.5V 검증.
        취소 시 None (위저드 중단)."""
        done = threading.Event()
        result = {"v": None}

        def ask():
            while True:
                s = simpledialog.askstring("DAC Calibration", prompt,
                                           parent=self.root)
                if s is None:
                    break                      # 취소
                s = s.strip().replace(",", ".")
                try:
                    v = float(s)
                except ValueError:
                    messagebox.showerror("DAC Calibration",
                                         f"숫자를 입력하세요 (예: 0.5023). 입력: {s!r}")
                    continue
                if not 0.0 <= v <= 3.5:
                    messagebox.showerror("DAC Calibration",
                                         f"상식 범위(0~3.5V)를 벗어났어요: {v} V\n"
                                         "단위가 V인지 확인하세요 (mV 아님)")
                    continue
                result["v"] = v
                break
            done.set()

        self.root.after(0, ask)
        done.wait()
        return result["v"]

    def gui_input(self, prompt: str = "") -> str:
        self.q.put(prompt + "\n")
        done = threading.Event()
        result = {"v": ""}

        def ask():
            if "y/n" in prompt.lower():
                result["v"] = "y" if messagebox.askyesno("CHIP1", prompt) else "n"
            else:
                messagebox.showinfo("CHIP1", prompt or "Continue?")
            done.set()

        self.root.after(0, ask)
        done.wait()
        self.q.put(f"  -> {result['v'] or '(ok)'}\n")
        return result["v"]

    # ---------------- run ----------------

    def validate(self) -> dict:
        def uv(s: str, name: str) -> int:
            try:
                v = int(s.strip())
            except ValueError:
                raise ValueError(f"{name}: 정수 µV 값이어야 해요")
            if not 0 <= v <= 3300000:
                raise ValueError(f"{name}: 0~3300000 µV 범위여야 해요 (VDDA=3.3V)")
            return v

        ainp = uv(self.ainp.get(), "AINP (dac set 1)")
        ainn = uv(self.ainn.get(), "AINN (dac set 2)")

        try:
            n = int(self.nsamp.get().strip())
        except ValueError:
            raise ValueError("'How many samples'는 정수여야 해요")
        if n <= 0:
            raise ValueError("'How many samples'는 1 이상이어야 해요")

        try:
            settle = float(self.settle.get().strip() or "0")
        except ValueError:
            raise ValueError("'Settle wait sec'는 숫자여야 해요")

        modes = [(key, reg) for key, _, reg in MODES if self.mode_vars[key].get()]
        if not modes:
            raise ValueError("모드를 하나 이상 선택하세요")

        dut_s = self.dut.get().strip()
        dut = int(dut_s) if dut_s else None
        return dict(ainp=ainp, ainn=ainn, n=n, settle=settle, modes=modes,
                    dut=dut, excel=Path(self.excel.get().strip()),
                    port=norm_port(self.port.get()) or None)

    # ---------------- ADC(meas) calibration wizard ----------------

    @staticmethod
    def _meas_cal_summary() -> str:
        """meas_cal.json 요약: 셀프캘 이력 + (Advanced) 외부 2점 계수."""
        import json
        try:
            d = (json.loads(g.MEAS_CAL_FILE.read_text(encoding="utf-8"))
                 if g.MEAS_CAL_FILE.exists() else {})
        except (ValueError, OSError):
            d = {}
        sc = d.get("selfcal")
        if sc:
            head = (f"Self-cal: {sc.get('date', '?')[:16]} "
                    f"VDDA={sc.get('vdda_v', '?')}V"
                    + (f" 검증오차 {sc['verify_err_mv']}mV {sc.get('verify', '')}"
                       if 'verify_err_mv' in sc else " (DMM 검증 전)"))
        else:
            head = "Self-cal: 미실행"
        ext = [f"{n}: 외부캘 {d[k].get('date', '?')[:10]}"
               for k, n in (("vref", "VREF"), ("vdd", "VDD")) if d.get(k)]
        return head + ("  |  " + "  |  ".join(ext) if ext else "")

    def self_cal_clicked(self):
        if not messagebox.askokcancel(
                "ADC Self-cal",
                "내부 셀프캘을 실행합니다 (배선 변경 불필요, 수 초 소요)\n\n"
                "ADCAL(오프셋) 재실행 + VREFINT 공장캘 재기준(게인).\n"
                "실행 후 DMM으로 AVDD 구멍 실측값을 입력하면\n"
                "±5mV 기준으로 검증까지 기록됩니다 (건너뛰기 가능)."):
            return
        self.selfcal_btn.configure(state="disabled")
        self.acal_btn.configure(state="disabled")
        self.run_btn.configure(state="disabled")
        threading.Thread(target=self._self_cal_worker, daemon=True).start()

    def _self_cal_worker(self):
        old = (sys.stdout, sys.stderr)
        sys.stdout = sys.stderr = QueueWriter(self.q)
        try:
            print("=== ADC self-cal ===")
            port = norm_port(self.port.get()) or g.find_port()
            with g.open_port(port) as ser:
                info = g.meas_selfcal(ser)
                print(f"  CALFACTOR={info['calfactor']} "
                      f"VREFINT_RAW={info['vrefint_raw']} VDDA={info['vdda_v']}V")
                vdd_v, vref_v = self._parse_meas(g.send_cmd(ser, "meas"))
                print(f"  meas: VDD={vdd_v}V VREF={vref_v}V")

            dmm = self.gui_ask_volts(
                f"셀프캘 완료. meas VREF = {vref_v}V\n\n"
                "검증: DMM으로 AVDD 구멍(개조 보드) 실측값을 V 단위로 입력\n"
                "(취소 = 검증 생략, 셀프캘 기록만 저장)")
            if dmm is not None and vref_v is not None:
                err_mv = (vref_v - dmm) * 1000.0
                passed = abs(err_mv) <= 5.0
                verdict = "PASS" if passed else "FAIL"
                print(f"  검증: meas {vref_v}V vs DMM {dmm}V "
                      f"-> 차이 {err_mv:+.1f}mV {verdict} (기준 ±5mV)")
                g.save_selfcal_record(info, round(err_mv, 2), verdict)
                msg = (f"Self-cal {verdict}\nVDDA={info['vdda_v']}V\n"
                       f"meas-DMM 차이 = {err_mv:+.1f}mV (기준 ±5mV)")
                if not passed:
                    msg += "\n\nFAIL — 배선/접촉 확인 후 Advanced(외부 2점) 캘 권장"
            else:
                g.save_selfcal_record(info)
                msg = f"Self-cal 기록됨 (DMM 검증 생략)\nVDDA={info['vdda_v']}V"
            self.root.after(0, lambda m=msg: messagebox.showinfo("ADC Self-cal", m))
        except g.serial.SerialException as e:
            print(f"\n!! 시리얼 포트 오류: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "ADC Self-cal", "COM 포트를 열 수 없습니다.\n"
                "Tera Term 등 다른 프로그램을 닫고 다시 시도하세요."))
        except (g.CliError, SystemExit) as e:
            print(f"\n!! self-cal 실패: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "ADC Self-cal", f"셀프캘 실패:\n{e}\n(구펌웨어면 재플래시 필요 — "
                "meas cal 명령은 2026-07-23 펌웨어부터)"))
        except Exception:
            traceback.print_exc()
        finally:
            sys.stdout, sys.stderr = old
            self.root.after(0, lambda: (
                self.selfcal_btn.configure(state="normal"),
                self.acal_btn.configure(state="normal"),
                self.run_btn.configure(state="normal", text="Run test"),
                self.acal_status.set(self._meas_cal_summary())))

    def adc_cal_clicked(self):
        chkey = "vref" if self.acal_ch.get().startswith("VREF") else "vdd"
        pin = "A4 (CN8-5)" if chkey == "vref" else "A3 (CN8-4)"
        try:
            low = int(self.acal_low.get().strip())
            high = int(self.acal_high.get().strip())
        except ValueError:
            messagebox.showerror("ADC Calibration", "저점/고점은 정수 µV여야 해요")
            return
        if not (0 <= low <= 3300000 and 0 <= high <= 3300000) or high - low < 500000:
            messagebox.showerror("ADC Calibration",
                                 "0~3300000 µV 범위 + 간격 0.5V 이상이어야 해요")
            return
        if not messagebox.askokcancel(
                "ADC Calibration",
                f"meas {chkey.upper()} 채널 2점 캘 시작\n\n"
                f"⚠ 임시 배선 필요: {pin}에서 원래 선을 빼고,\n"
                f"   D1(PA4, DAC ch1 출력)을 {pin}에 연결하세요.\n"
                f"   (DAC ch1은 캘 완료 상태라 기준 소스로 사용)\n\n"
                f"저점 {low / 1e6:.3f}V / 고점 {high / 1e6:.3f}V, "
                f"검증 {CAL_VERIFY_UV / 1e6:.3f}V (±{CAL_TOL_UV / 1000:.0f}mV)\n"
                "배선 완료 후 확인을 누르세요."):
            return
        self.acal_btn.configure(state="disabled")
        self.run_btn.configure(state="disabled")
        threading.Thread(target=self._adc_cal_worker, args=(chkey, pin, low, high),
                         daemon=True).start()

    def _adc_cal_worker(self, chkey: str, pin: str, low: int, high: int):
        old = (sys.stdout, sys.stderr)
        sys.stdout = sys.stderr = QueueWriter(self.q)

        def read_ch(ser) -> float:
            vdd_v, vref_v = self._parse_meas(g.send_cmd(ser, "meas"))
            v = vref_v if chkey == "vref" else vdd_v
            if v is None:
                raise g.CliError("meas 응답 파싱 실패")
            return v

        try:
            print(f"=== ADC(meas) 2-point cal: {chkey}  low={low} high={high} ===")
            port = norm_port(self.port.get()) or g.find_port()
            with g.open_port(port) as ser:
                g.dac_ready(ser)   # 캘 적용된 DAC ch1이 기준 소스

                g.send_cmd(ser, f"dac set 1 {low}")
                time.sleep(0.3)
                m1 = read_ch(ser)
                print(f"  저점: 지시 {low / 1e6:.3f}V -> meas {m1:.3f}V")

                g.send_cmd(ser, f"dac set 1 {high}")
                time.sleep(0.3)
                m2 = read_ch(ser)
                print(f"  고점: 지시 {high / 1e6:.3f}V -> meas {m2:.3f}V")

                offset_uv, gain_ppm = cal_two_point(low, m1 * 1e6, high, m2 * 1e6)
                print(f"  계수: offset_uV={offset_uv}, gain_ppm={gain_ppm}")

                g.send_cmd(ser, f"dac set 1 {CAL_VERIFY_UV}")
                time.sleep(0.3)
                m3 = read_ch(ser)
                corrected = (m3 * 1e6 + offset_uv) * (1.0 + gain_ppm / 1e6)
                err_uv = corrected - CAL_VERIFY_UV
                passed = abs(err_uv) <= CAL_TOL_UV
                verdict = "PASS" if passed else "FAIL"
                print(f"  검증: 지시 {CAL_VERIFY_UV / 1e6:.3f}V, 보정 후 "
                      f"{corrected / 1e6:.4f}V, 오차 {err_uv / 1000:+.2f}mV -> {verdict}")

                save = passed or messagebox.askyesno(
                    "ADC Calibration",
                    f"검증 FAIL (오차 {err_uv / 1000:+.2f}mV). 그래도 저장할까요?")
                if save:
                    g.save_meas_cal(chkey, offset_uv, gain_ppm,
                                    extra={"verify_err_mv": round(err_uv / 1000, 3),
                                           "verify": verdict,
                                           "points_uv": [low, high]})
                    print(f"  저장됨: {g.MEAS_CAL_FILE.name}")
                g.send_cmd(ser, "dac set 1 0")
            self.root.after(0, lambda: messagebox.showinfo(
                "ADC Calibration",
                f"meas {chkey.upper()} 캘 {verdict}\n\n"
                f"offset = {offset_uv} µV / gain = {gain_ppm} ppm\n"
                f"검증 오차 = {err_uv / 1000:+.2f} mV\n\n"
                f"⚠ 임시 배선을 제거하고 {pin}을 원래 노드로 복구하세요!"))
        except g.serial.SerialException as e:
            print(f"\n!! 시리얼 포트 오류: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "ADC Calibration", "COM 포트를 열 수 없습니다.\n"
                "Tera Term 등 다른 프로그램을 닫고 다시 시도하세요."))
        except (g.CliError, ValueError, SystemExit) as e:
            print(f"\n!! ADC cal 실패: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "ADC Calibration", f"캘리브레이션 실패:\n{e}"))
        except Exception:
            traceback.print_exc()
        finally:
            sys.stdout, sys.stderr = old
            self.root.after(0, lambda: (
                self.acal_btn.configure(state="normal"),
                self.run_btn.configure(state="normal", text="Run test"),
                self.acal_status.set(self._meas_cal_summary())))

    # ---------------- calibration wizard ----------------

    def cal_clicked(self):
        try:
            ch = int(self.cal_ch.get())
            low = int(self.cal_low.get().strip())
            high = int(self.cal_high.get().strip())
        except ValueError:
            messagebox.showerror("DAC Calibration", "저점/고점은 정수 µV여야 해요")
            return
        if ch not in (1, 2):
            messagebox.showerror("DAC Calibration", "채널은 1 또는 2")
            return
        if not (0 <= low <= 3300000 and 0 <= high <= 3300000):
            messagebox.showerror("DAC Calibration", "저점/고점은 0~3300000 µV 범위")
            return
        if high - low < 500000:
            messagebox.showerror("DAC Calibration",
                                 "고점-저점 간격이 0.5V 이상이어야 기울기가 정확해요")
            return
        pin = "D1 (PA4)" if ch == 1 else "D13 (PA5)"
        if not messagebox.askokcancel(
                "DAC Calibration",
                f"ch{ch} 캘리브레이션 시작\n\n"
                f"멀티미터: 빨강 → {pin}, 검정 → GND\n"
                f"저점 {low / 1e6:.3f}V / 고점 {high / 1e6:.3f}V, "
                f"검증 {CAL_VERIFY_UV / 1e6:.3f}V (±{CAL_TOL_UV / 1000:.0f}mV)\n\n"
                "준비되면 확인을 누르세요."):
            return
        self.cal_btn.configure(state="disabled")
        self.run_btn.configure(state="disabled")
        threading.Thread(target=self._cal_worker, args=(ch, low, high),
                         daemon=True).start()

    def _cal_worker(self, ch: int, low: int, high: int):
        old = (sys.stdout, sys.stderr)
        sys.stdout = sys.stderr = QueueWriter(self.q)
        try:
            print(f"=== 2-point cal: ch{ch}  low={low} high={high} ===")
            port = self.port.get().strip() or g.find_port()
            with g.open_port(port) as ser:
                # 기존 보정을 0으로 리셋한 '날것' 상태에서 2점을 측정해야
                # 전달함수 m = A*s + B 를 올바르게 추정할 수 있다
                g.send_cmd(ser, "dac init")
                g.send_cmd(ser, f"dac cal {ch} 0 0")

                g.send_cmd(ser, f"dac set {ch} {low}")
                m1 = self.gui_ask_volts(
                    f"[1/3] 저점 출력 중: 지시 {low / 1e6:.3f}V\n"
                    "멀티미터 실측값을 V 단위로 입력:")
                if m1 is None:
                    print("  취소됨")
                    return

                g.send_cmd(ser, f"dac set {ch} {high}")
                m2 = self.gui_ask_volts(
                    f"[2/3] 고점 출력 중: 지시 {high / 1e6:.3f}V\n"
                    "멀티미터 실측값을 V 단위로 입력:")
                if m2 is None:
                    print("  취소됨")
                    return

                offset_uv, gain_ppm = cal_two_point(
                    low, m1 * 1e6, high, m2 * 1e6)
                print(f"  계수 산출: offset_uV={offset_uv}, gain_ppm={gain_ppm}")
                if abs(gain_ppm) > 50000 or abs(offset_uv) > 100000:
                    print("  !! 계수가 비정상적으로 큼 (gain > 5% 또는 offset > 100mV)")
                    if not messagebox.askyesno(
                            "DAC Calibration",
                            f"계수가 비정상적으로 커요:\n"
                            f"offset={offset_uv}µV, gain={gain_ppm}ppm\n"
                            "실측 입력이 맞나요? 계속할까요?"):
                        return

                # 새 계수 적용 → 중간점 검증 (cal은 다음 set부터 반영되므로 순서 중요)
                g.send_cmd(ser, f"dac cal {ch} {offset_uv} {gain_ppm}")
                g.send_cmd(ser, f"dac set {ch} {CAL_VERIFY_UV}")
                m3 = self.gui_ask_volts(
                    f"[3/3] 검증점 출력 중: 지시 {CAL_VERIFY_UV / 1e6:.3f}V (보정 적용됨)\n"
                    "멀티미터 실측값을 V 단위로 입력:")
                if m3 is None:
                    print("  취소됨 (계수는 저장 안 함)")
                    return

                err_uv = m3 * 1e6 - CAL_VERIFY_UV
                passed = abs(err_uv) <= CAL_TOL_UV
                verdict = "PASS" if passed else "FAIL"
                print(f"  검증: 지시 {CAL_VERIFY_UV / 1e6:.3f}V, 실측 {m3:.4f}V, "
                      f"오차 {err_uv / 1000:+.2f}mV → {verdict}")

                save = passed or messagebox.askyesno(
                    "DAC Calibration",
                    f"검증 FAIL (오차 {err_uv / 1000:+.2f}mV > "
                    f"±{CAL_TOL_UV / 1000:.0f}mV).\n그래도 계수를 저장할까요?")
                board_saved = False
                if save:
                    g.save_dac_cal(ch, offset_uv, gain_ppm,
                                   extra={"verify_err_mv": round(err_uv / 1000, 3),
                                          "verify": verdict,
                                          "points_uv": [low, high]})
                    print(f"  저장됨: {g.CAL_FILE.name} (Run test 시 자동 적용)")
                    # 보드 플래시에 **이 채널만** 영구 저장 (반대 채널 저장분
                    # 보존 — 채널별 valid 플래그). + PC 레지스트리에 UID 기록.
                    other_missing = False
                    try:
                        if g.send_cmd(ser, f"dac cal save {ch}") is not None:
                            board_saved = True
                            uid = g.read_uid(ser)
                            fch = g.parse_flash_state(
                                g.send_cmd(ser, "dac cal show"))
                            g.registry_note(uid, flash_ch1=fch[1],
                                            flash_ch2=fch[2],
                                            last_cal=f"ch{ch} {verdict}",
                                            last_cal_err_mv=round(err_uv / 1000, 3))
                            other = 2 if ch == 1 else 1
                            other_missing = not fch[other]
                            print(f"  보드 플래시 저장 완료: ch{ch} (UID {uid})")
                    except g.CliError as e:
                        print(f"  !! 보드 플래시 저장 실패({e}) — json 계수는 유효")
                def _finish(bs=board_saved, om=other_missing, cur=ch):
                    messagebox.showinfo(
                        "DAC Calibration",
                        f"ch{cur} 캘리브레이션 {verdict}\n\n"
                        f"offset = {offset_uv} µV\ngain = {gain_ppm} ppm\n"
                        f"검증 오차 = {err_uv / 1000:+.2f} mV (목표 ±2mV)\n\n"
                        + ("보드 플래시 + dac_cal.json에 저장됨 — 이 보드는 이제\n"
                           "어느 PC에 꽂아도 보정이 유지됩니다."
                           if bs else "dac_cal.json에 저장됨 (보드 저장은 안 됨)."
                           if save else "저장하지 않았습니다."))
                    # 체이닝: 반대 채널이 아직 미캘이면 이어서 진행 제안
                    if bs and om:
                        other = 2 if cur == 1 else 1
                        pin_o = "D1" if other == 1 else "D13"
                        if messagebox.askyesno(
                                "DAC Calibration",
                                f"ch{other}는 아직 미캘 상태입니다.\n"
                                f"이어서 ch{other} 캘리브레이션을 진행할까요?\n"
                                f"(멀티미터를 {pin_o}로 옮겨 물리면 됩니다)"):
                            self.cal_ch.set(str(other))
                            # finally의 버튼 재활성화가 먼저 처리되도록 지연 후
                            # 시작 (즉시 호출하면 2차 위저드 중 버튼이 풀리는
                            # 레이스 — 2026-07-24 리뷰에서 선제 수정)
                            self.root.after(300, self.cal_clicked)
                self.root.after(0, _finish)
        except g.serial.SerialException as e:
            print(f"\n!! 시리얼 포트 오류: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "DAC Calibration",
                "COM 포트를 열 수 없습니다.\n\n"
                "Tera Term 등 다른 프로그램이 포트를 잡고 있으면\n"
                "닫은(Disconnect) 다음 다시 시도하세요."))
        except (g.CliError, ValueError, SystemExit) as e:
            print(f"\n!! cal 실패: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "DAC Calibration", f"캘리브레이션 실패:\n{e}"))
        except Exception:
            traceback.print_exc()
        finally:
            sys.stdout, sys.stderr = old
            self.root.after(0, lambda: (
                self.cal_btn.configure(state="normal"),
                self.run_btn.configure(state="normal", text="Run test")))

    # ---------------- temperature sweep ----------------

    def sweep_clicked(self):
        try:
            temps = [float(t) for t in self.sw_temps.get().replace(" ", "").split(",") if t]
            if not temps:
                raise ValueError("온도 리스트가 비었어요")
            soak = float(self.sw_soak.get())
            tol = float(self.sw_tol.get())
            if soak < 0 or tol <= 0:
                raise ValueError("포화시간 >= 0, 허용오차 > 0 이어야 해요")
            base = self.validate()          # 기존 Run test 입력 재검증 (전압/샘플수 등)
        except ValueError as e:
            messagebox.showerror("Temp Sweep", str(e))
            return
        profile = self.sw_profile.get()
        cport = norm_port(self.sw_port.get())
        if profile != "MOCK" and not cport:
            messagebox.showerror("Temp Sweep", "챔버 포트를 입력하세요 (MOCK 제외)")
            return
        est_min = len(temps) * (soak + base["n"] / 10 / 60 * 2 + 5)
        if not messagebox.askokcancel(
                "Temp Sweep",
                f"{len(temps)}개 온도 스윕 시작\n{temps}\n\n"
                f"포화 {soak:g}분, 허용오차 ±{tol:g}°C, 챔버 {profile}\n"
                f"대략 소요 {est_min:.0f}분 이상 (도달 시간 제외)\n\n"
                "보드/챔버 연결 확인 후 확인을 누르세요."):
            return
        self.sweep_btn.configure(state="disabled")
        self.run_btn.configure(state="disabled")
        self.cal_btn.configure(state="disabled")
        self.sweep_stop_btn.configure(state="normal")
        params = dict(base, temps=temps, soak=soak, tol=tol,
                      profile=profile, cport=cport,
                      swlabel=self.sw_label.get().strip(),
                      end_mode=self.sw_end.get(),
                      selfcal=self.sw_selfcal.get())
        threading.Thread(target=self._sweep_worker, args=(params,),
                         daemon=True).start()

    def sweep_stop_clicked(self):
        if not getattr(self, "_sweep", None):
            return
        ans = messagebox.askyesnocancel(
            "Temp Sweep 중단",
            "어떻게 중단할까요?\n\n"
            "예 = 즉시 중단 (현재 대기/측정 끊고 종료)\n"
            "아니오 = 현재 온도 스텝(기록까지) 마치고 중단\n"
            "취소 = 계속 진행")
        if ans is None:
            return
        self._sweep.request_stop(immediate=bool(ans))
        self.q.put(f"\n[sweep] 중단 요청됨 ({'즉시' if ans else '현재 스텝 후'})\n")

    def _sweep_status(self, st):
        """상태머신 → GUI 라벨/로그 (워커 스레드에서 호출됨)."""
        soak_m, soak_s = divmod(int(st.soak_remaining_s), 60)
        text = (f"[{st.state}] 온도 {st.temp_idx + 1 if st.state != 'DONE' else st.temp_total}"
                f"/{st.temp_total}  목표 {st.set_temp:g}°C"
                f"  챔버 {st.chamber_temp if st.chamber_temp is not None else '?'}°C"
                + (f"  포화 남음 {soak_m}:{soak_s:02d}" if st.state == "SOAK" else "")
                + (f"  — {st.message}" if st.message else ""))
        self.root.after(0, lambda: self.sw_status.set(text))
        if st.message and st.state in ("SET_TEMP", "TEST", "REPORT",
                                       "DONE", "ABORTED", "ERROR"):
            self.q.put(f"[sweep] {text}\n")

    @staticmethod
    def _parse_meas(payload_lines):
        """'VDD=3.299V VREF=3.001V' → (3.299, 3.001). 실패 시 (None, None)."""
        for line in payload_lines:
            m = re.search(r"VDD=([\d.]+)V\s+VREF=([\d.]+)V", line)
            if m:
                return float(m.group(1)), float(m.group(2))
        return None, None

    def _resolve_next_dut(self, excel: Path) -> int:
        """다음 빈 DUT 블록 번호 (없으면 자동 확장)."""
        wb = self._load_report(excel)
        ws = wb[g.SHEET_NAME] if g.SHEET_NAME else wb.active
        dut = g.find_next_empty_dut(ws)
        duts = _scan_duts(ws)
        wb.close()
        if dut is None:
            if not duts:
                raise SystemExit("리포트에서 DUT# 헤더를 못 찾았어요")
            dut = max(duts) + 1
        self._ensure_dut_block(excel, dut)
        return dut

    def _label_block_temp(self, excel: Path, dut: int, temp_c: float,
                          run_tag: str = ""):
        """DUT 블록 헤더에 온도+런 라벨: '... DUT#7 @-40C [0723_1706]'
        런 태그 = results/sweeps/<런폴더> 역추적용."""
        wb = self._load_report(excel)
        ws = wb[g.SHEET_NAME] if g.SHEET_NAME else wb.active
        ci, _ = g.locate_dut_columns(ws, dut)
        cell = ws.cell(row=g.HEADER_ROW, column=ci)
        base = re.sub(r"\s*@.*$", "", str(cell.value or f"DUT#{dut}"))
        tag = f" [{run_tag}]" if run_tag else ""
        cell.value = f"{base} @{temp_c:g}C{tag}"
        while True:
            try:
                wb.save(excel)
                return
            except PermissionError:
                input("엑셀 파일이 열려 있어요! Excel을 닫은 다음 OK를 눌러 주세요.")

    def _sweep_test_one(self, temp_c: float, p: dict, csv_writer, csv_file):
        """한 온도에서: meas → 기존 테스트 로직 → 엑셀(온도 라벨) + CSV 행."""
        excel = p["excel"]
        rows = []
        port = p["port"] or g.find_port()
        with g.open_port(port) as ser:
            g.check_id(ser)

            g.dac_ready(ser)   # 보드 플래시 캘 우선, 없으면 json 폴백
            g.send_cmd(ser, f"dac set 1 {p['ainp']}")
            g.send_cmd(ser, f"dac set 2 {p['ainn']}")

            # 온도별 ADC 셀프캘 (ADCAL+VREFINT 재기준) — 온도 드리프트 대응
            if p.get("selfcal"):
                try:
                    info = g.meas_selfcal(ser)
                    print(f"  self-cal @{temp_c:g}C: CALFACTOR={info['calfactor']} "
                          f"VDDA={info['vdda_v']}V")
                except g.CliError as e:
                    print(f"  !! self-cal 실패 ({e}) — 계속 진행")

            # 온도 스텝마다 meas 기록 (배선 전이면 경고만 하고 계속)
            try:
                vdd_v, vref_v = self._parse_meas(g.send_cmd(ser, "meas"))
                vdd_v, vref_v = g.apply_meas_cal(vdd_v, vref_v)   # meas_cal.json 보정
            except g.CliError as e:
                print(f"  !! meas 실패 ({e}) — 빈 값으로 기록")
                vdd_v = vref_v = None
            lo, hi = SWEEP_VREF_RANGE
            vref_valid = vref_v is not None and lo <= vref_v <= hi
            if not vref_valid:
                # 스윕은 중단하지 않는다 — VDD만으로도 유효한 데이터
                shown = f"{vref_v:.3f}V" if vref_v is not None else "측정 실패"
                print(f"  [주의] VREF 범위 밖 ({shown}, 기대 {lo}~{hi}V) — "
                      "미개조 보드이거나 칩 미장착/접촉 불량.\n"
                      "         CSV에는 측정값 그대로 기록 (vref_valid=False)")

            dut = self._resolve_next_dut(excel)
            print(f"  이 온도의 DUT 블록: #{dut} @{temp_c:g}C")
            chamber_temp = self._sweep.status.chamber_temp

            wr_history = []

            def wr_tracked(addr, val):
                g.write_verified(ser, addr, val)
                wr_history.append(f"wr {addr} {val}")

            n = p["n"]
            for mode, reg04 in p["modes"]:
                print(f"\n[{MODE_LABELS[mode]}] DUT #{dut} @{temp_c:g}C")
                wr_tracked("0x03", "0x02")
                wr_tracked("0x04", reg04)
                if p["settle"] > 0:
                    time.sleep(p["settle"])
                if g.SETTLE_RD > 0:
                    g.send_cmd(ser, f"rd {g.SETTLE_RD}",
                               timeout_s=g.RD_LINE_TIMEOUT, quiet=True)
                values = g.capture_samples(ser, n)
                csv_path = g.save_csv(values, dut, f"{mode}_{temp_c:g}C",
                                      out_dir=p["run_dir"] / "captures")
                self._write_excel_retry(values, dut, mode, excel)

                row = {
                    "timestamp": datetime.now().isoformat(timespec="seconds"),
                    "set_temp_C": temp_c,
                    "chamber_temp_C": chamber_temp,
                    "mode": mode,
                    "VDD_V": vdd_v, "VREF_V": vref_v,
                    "vref_valid": vref_valid,
                    "reg_settings": ";".join(wr_history[-3:]),   # p.5 2안
                    "rd_mean": round(statistics.fmean(values), 1),
                    "rd_stdev": round(statistics.stdev(values), 1) if len(values) > 1 else 0,
                    "rd_count": len(values),
                    "dut_block": dut,
                    "sample_csv": csv_path.name,
                }
                csv_writer.writerow(row)
                csv_file.flush()
                rows.append(row)

            self._label_block_temp(excel, dut, temp_c, p.get("run_tag", ""))
            self._write_meta_retry(excel, dut, vdd_v, vref_v,
                                   p["ainp"], p["ainn"])
        return rows

    def _sweep_worker(self, p: dict):
        start_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        run_log: list[str] = []
        old = (sys.stdout, sys.stderr, builtins.input)
        sys.stdout = sys.stderr = QueueWriter(self.q, run_log)
        builtins.input = self.gui_input
        chamber = None
        csv_file = None
        prev_sheet = g.SHEET_NAME
        try:
            g.ensure_validation_workbook(p["excel"], g.TEMPLATE_XLSX)
            print(f"=== temp sweep {start_ts}  temps={p['temps']} "
                  f"soak={p['soak']}min tol=±{p['tol']}C chamber={p['profile']} ===")

            # 스윕 전용 시트 생성 — 상온 검증 시트는 건드리지 않는다
            sweep_sheet = create_sweep_sheet(p["excel"], p.get("swlabel", ""))
            g.SHEET_NAME = sweep_sheet
            print(f"  스윕 시트 생성: [{sweep_sheet}] (블록 = 온도)")

            # 런 폴더: 스윕 1회 = results/sweeps/<타임스탬프>/ (재현성 단위)
            run_dir = g.SWEEPS_DIR / start_ts
            (run_dir / "captures").mkdir(parents=True, exist_ok=True)
            p["run_dir"] = run_dir
            p["run_tag"] = f"{start_ts[4:8]}_{start_ts[9:13]}"   # MMDD_HHMM
            (run_dir / "run_info.txt").write_text(
                "CHIP1 temperature sweep run\n"
                f"start        : {start_ts}\n"
                f"label        : {p.get('swlabel') or '(없음)'}\n"
                f"excel sheet  : {sweep_sheet}  (파일: {p['excel'].name})\n"
                f"temps_C      : {p['temps']}\n"
                f"soak_min     : {p['soak']}  / tol ±{p['tol']}C\n"
                f"chamber      : {p['profile']} @ {p.get('cport') or 'MOCK'}\n"
                f"end_mode     : {p.get('end_mode')}\n"
                f"samples/mode : {p['n']}  (settle {p['settle']}s + rd {g.SETTLE_RD} 폐기)\n"
                f"dac stimulus : AINP(ch1)={p['ainp']}uV, AINN(ch2)={p['ainn']}uV\n"
                f"dac_cal      : {g.load_dac_cal()}\n"
                f"meas_cal     : {g.load_meas_cal()}\n"
                f"modes        : {[m for m, _ in p['modes']]}\n",
                encoding="utf-8")
            print(f"  런 폴더: {run_dir.relative_to(g.PROJECT_ROOT)}")

            if p["profile"] == "MOCK":
                chamber = MockChamber(start_c=25.0, rate_c_per_s=2.0)
            else:
                chamber = Chamber(p["cport"], profile=p["profile"])
            chamber.start()   # POWER,ON + 습도 알람 한계 (C# 예제 시퀀스)

            report_path = run_dir / "sweep_report.csv"   # 폴더가 런을 식별
            csv_file = open(report_path, "w", newline="", encoding="utf-8-sig")
            csv_writer = csv.DictWriter(csv_file, fieldnames=SWEEP_CSV_FIELDS)
            csv_writer.writeheader()

            self._sweep = TempSweep(
                chamber,
                lambda t: self._sweep_test_one(t, p, csv_writer, csv_file),
                SweepConfig(temps=p["temps"], soak_min=p["soak"],
                            tol_c=p["tol"]),
                status_cb=self._sweep_status)
            st = self._sweep.run()

            print(f"\n=== sweep {st.state}: {st.message} ===")
            print(f"  리포트: {report_path.relative_to(g.PROJECT_ROOT)} ({len(st.rows)}행)")

            # 종료 동작 (챔버 방치 방지). 중단/에러 시에는 복귀 대기 없이 즉시 정지.
            end_mode = p.get("end_mode", "유지(수동)")
            try:
                if end_mode.startswith("25") and st.state == "DONE":
                    print("[sweep] 종료 동작: 25C 복귀 후 정지 (최대 60분 대기)")
                    chamber.set_temp(25)
                    deadline = time.time() + 3600
                    while time.time() < deadline:
                        cur, _h, _s = chamber.get_mon()
                        self.q.put(f"[sweep] 상온 복귀 중... {cur}C\n")
                        if abs(cur - 25) <= 2.0:
                            break
                        time.sleep(10)
                    chamber.power(False)
                    print("[sweep] 챔버 정지 (POWER,OFF)")
                elif end_mode != "유지(수동)":
                    chamber.power(False)
                    print("[sweep] 챔버 즉시 정지 (POWER,OFF)")
                else:
                    print("[sweep] 챔버 유지 — 수동으로 정지/복귀 필요!")
            except Exception as e:                     # noqa: BLE001
                print(f"!! 종료 동작 실패: {e} — 챔버 상태 수동 확인 필요!")
            self.root.after(0, lambda: messagebox.showinfo(
                "Temp Sweep",
                f"스윕 {st.state}\n{st.message}\n\n"
                f"완료 온도: {st.temp_idx}/{st.temp_total}\n"
                f"런 폴더: results/sweeps/{start_ts}/ (리포트+캡처+조건)\n"
                f"엑셀: {p['excel'].name} 의 [{sweep_sheet}] 시트 (블록 = 온도)"))
        except g.serial.SerialException as e:
            print(f"\n!! 시리얼 포트 오류: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "Temp Sweep",
                "COM 포트를 열 수 없습니다 (보드 또는 챔버).\n\n"
                "Tera Term 등 다른 프로그램이 포트를 잡고 있으면\n"
                "닫은 다음 다시 시도하세요."))
        except (ChamberError, SystemExit, g.CliError) as e:
            print(f"\n!! sweep 실패: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "Temp Sweep", f"스윕 실패:\n{e}"))
        except Exception:
            traceback.print_exc()
        finally:
            if chamber is not None:
                try:
                    chamber.close()
                except Exception:
                    pass
            if csv_file is not None:
                csv_file.close()
            self._sweep = None
            g.SHEET_NAME = prev_sheet   # 스윕 시트 지정 해제 (상온 테스트 복귀)
            sys.stdout, sys.stderr, builtins.input = old
            try:
                log_dir = p.get("run_dir") or g.SWEEPS_DIR
                log_dir.mkdir(parents=True, exist_ok=True)
                (log_dir / f"sweep_{start_ts}.log").write_text(
                    "".join(run_log), encoding="utf-8")
            except OSError:
                pass
            self.root.after(0, lambda: (
                self.sweep_btn.configure(state="normal"),
                self.run_btn.configure(state="normal", text="Run test"),
                self.cal_btn.configure(state="normal"),
                self.sweep_stop_btn.configure(state="disabled")))

    def run_clicked(self):
        try:
            params = self.validate()
        except ValueError as e:
            messagebox.showerror("CHIP1", str(e))
            return
        if params["n"] != g.N_SAMPLES and not messagebox.askyesno(
                "CHIP1",
                f"엑셀 리포트 수식은 샘플 {g.N_SAMPLES}개 기준입니다.\n"
                f"{params['n']}개로 진행하면 ENOB/STD 수식이 맞지 않을 수 "
                "있어요.\n계속할까요?"):
            return
        self._stop_run = False
        self.run_btn.configure(state="disabled", text="Running…  (see log)")
        self.run_stop_btn.configure(state="normal")
        threading.Thread(target=self._worker, args=(params,), daemon=True).start()

    def run_stop_clicked(self):
        """Run test 중단: 플래그 + 진행 중인 rd에 ESC 주입 (즉시 끊김)."""
        self._stop_run = True
        ser = self._active_ser
        if ser is not None:
            try:
                ser.write(b"\x1b")
            except Exception:
                pass
        self.q.put("\n[중단] Run 중단 요청 — 현재 명령 정리 중...\n")

    def _worker(self, p: dict):
        start_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        run_log: list[str] = []
        old = (sys.stdout, sys.stderr, builtins.input)
        sys.stdout = sys.stderr = QueueWriter(self.q, run_log)
        builtins.input = self.gui_input
        try:
            print(f"=== run {start_ts}  AINP={p['ainp']} AINN={p['ainn']} "
                  f"n={p['n']} excel={p['excel'].name} ===")
            dut, summary = self._run_test(p)
            print("\n=== DONE ===")
            if p["excel"].exists():
                print(f"  opening {p['excel'].name} ...")
                os.startfile(p["excel"].resolve())
            result_txt = ("\n".join(summary) + "\n\n") if summary else ""
            self.root.after(0, lambda: messagebox.showinfo(
                "CHIP1",
                f"DUT #{dut} 완료!\n\n" + result_txt +
                "다음 칩 테스트 순서:\n"
                "  1) 보드 전원(USB) 분리\n"
                "  2) 칩 교체\n"
                "  3) USB 재연결 (cal은 다음 Run에서 자동 재입력)\n"
                "  4) Run test 다시 클릭\n\n"
                "※ 자동으로 열린 엑셀은 다음 테스트 전에 닫아 주세요"))
        except SystemExit as e:
            print(f"\n!! aborted: {e}")
        except g.serial.SerialException as e:
            print(f"\n!! 시리얼 포트 오류: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "CHIP1", "COM 포트를 열 수 없습니다.\n\nTera Term 등 다른 "
                "프로그램이 포트를 잡고 있으면 닫은 다음 다시 시도하세요."))
        except g.CliError as e:
            # Run 중단 버튼의 ESC가 rd를 끊으면 'ERR aborted'로 도착 -> 정상 중단
            if self._stop_run:
                print("\n== Run 중단됨 (사용자 요청) — 부분 데이터는 저장 안 됨 ==")
            else:
                print(f"\n!! CLI error: {e}")
        except Exception:
            traceback.print_exc()
        finally:
            self._active_ser = None
            self.root.after(0, lambda: self.run_stop_btn.configure(
                state="disabled"))
            sys.stdout, sys.stderr, builtins.input = old
            try:
                log_path = g.PROJECT_ROOT / f"run_{start_ts}.log"
                log_path.write_text("".join(run_log), encoding="utf-8")
                self.q.put(f"  run log saved: {log_path.name}\n")
            except OSError as e:
                self.q.put(f"  !! could not save run log: {e}\n")
            self.root.after(0, lambda: self.run_btn.configure(
                state="normal", text="Run test"))

    # ---------------- excel helpers ----------------

    def _load_report(self, excel: Path):
        from openpyxl import load_workbook
        while True:
            try:
                return load_workbook(excel)
            except PermissionError:
                input("엑셀 파일이 열려 있어요! Excel을 닫은 다음 OK를 눌러 주세요.")

    def _ensure_dut_block(self, excel: Path, dut: int):
        wb = self._load_report(excel)
        ws = wb[g.SHEET_NAME] if g.SHEET_NAME else wb.active
        duts = _scan_duts(ws)
        if not duts:
            raise SystemExit(f"row {g.HEADER_ROW}에서 DUT# 헤더를 못 찾았어요")
        if dut in duts:
            wb.close()
            return
        last = max(duts)
        if dut < last:
            raise SystemExit(f"DUT#{dut} 블록이 없는데 리포트에는 DUT#{last}까지 "
                             "있어요 — DUT 번호를 확인해 주세요")
        last_data_row = g.DATA_START_ROW + g.N_SAMPLES - 1
        for nd in range(last + 1, dut + 1):
            dst = duts[last] + 2 * (nd - last)
            _copy_block(ws, duts[last], dst, nd)
            leftover = sum(
                1 for r in range(g.DATA_START_ROW, last_data_row + 1)
                for c in (dst, dst + 1)
                if ws.cell(row=r, column=c).value is not None)
            if leftover:
                if input(f"새 DUT#{nd} 블록 자리에 기존 값 {leftover}개가 있어요.\n"
                         "지우고 계속할까요? [y/N] ").strip().lower() != "y":
                    raise SystemExit(f"DUT#{nd} 블록 생성 취소")
                for r in range(g.DATA_START_ROW, last_data_row + 1):
                    for c in (dst, dst + 1):
                        ws.cell(row=r, column=c).value = None
            print(f"  리포트에 DUT#{nd} 블록을 새로 만들었어요 (템플릿 복사)")
        while True:
            try:
                wb.save(excel)
                return
            except PermissionError:
                input("엑셀 파일이 열려 있어요! Excel을 닫은 다음 OK를 눌러 주세요.")

    def _write_excel_retry(self, values, dut, mode, excel):
        while True:
            try:
                g.write_excel(values, dut, mode, excel)
                return
            except PermissionError:
                input("엑셀 파일이 열려 있어서 저장할 수 없어요!\n"
                      "Excel을 닫은 다음 OK를 눌러 주세요 — 다시 저장합니다.")

    def _write_meta_retry(self, excel, dut, vdd_v, vref_v, ainp_uv, ainn_uv):
        """블록 메타(실측 VDD/AVDD + 실지령 자극) 기입 — 잠금 시 재시도."""
        while True:
            try:
                g.write_block_meta(excel, dut, vdd_v, vref_v, ainp_uv, ainn_uv)
                return
            except PermissionError:
                input("엑셀 파일이 열려 있어요! Excel을 닫은 다음 OK를 눌러 주세요.")

    # ---------------- test core ----------------

    def _run_test(self, p: dict):
        excel = p["excel"]
        # 검증용 사본이 없으면 원본에서 자동 생성 (원본은 불변)
        g.ensure_validation_workbook(excel, g.TEMPLATE_XLSX)

        dut = p["dut"]
        if dut is None:
            wb = self._load_report(excel)
            ws = wb[g.SHEET_NAME] if g.SHEET_NAME else wb.active
            dut = g.find_next_empty_dut(ws)
            duts = _scan_duts(ws)
            wb.close()
            if dut is None:
                if not duts:
                    raise SystemExit("리포트에서 DUT# 헤더를 못 찾았어요")
                dut = max(duts) + 1
                print(f"모든 DUT 블록이 가득 참 → 새 블록 DUT#{dut} 생성")
            else:
                print(f"auto-detected next empty DUT block: DUT#{dut}")
        self._ensure_dut_block(excel, dut)
        summary: list[str] = []

        def abort_check():
            if self._stop_run:
                raise SystemExit("사용자 중단 (Run 중단 버튼)")

        port = p["port"] or g.find_port()
        print(f"  port: {port}")
        with g.open_port(port) as ser:
            self._active_ser = ser        # Run 중단 버튼의 ESC 주입 대상
            g.check_id(ser)

            print("\n[internal DAC setup]")
            g.dac_ready(ser)   # 보드 플래시 캘 우선, 없으면 json 폴백
            g.send_cmd(ser, f"dac set 1 {p['ainp']}")   # ch1(PA4) → AINP
            g.send_cmd(ser, f"dac set 2 {p['ainn']}")   # ch2(PA5) → AINN
            abort_check()

            try:
                vdd_v, vref_v = g.apply_meas_cal(
                    *self._parse_meas(g.send_cmd(ser, "meas")))
                print(f"  meas: VDD={vdd_v}V VREF={vref_v}V")
            except g.CliError:
                vdd_v = vref_v = None

            n = p["n"]
            for mode, reg04 in p["modes"]:
                abort_check()
                print(f"\n[{MODE_LABELS[mode]}] DUT #{dut}")
                g.write_verified(ser, "0x03", "0x02")
                g.write_verified(ser, "0x04", reg04)

                if p["settle"] > 0:
                    print(f"  settling wait {p['settle']:.0f}s")
                    end = time.time() + p["settle"]
                    while time.time() < end:      # 중단 반응형 대기
                        abort_check()
                        time.sleep(min(0.5, max(0, end - time.time())))
                if g.SETTLE_RD > 0:
                    print(f"  discarding {g.SETTLE_RD} settling samples")
                    g.send_cmd(ser, f"rd {g.SETTLE_RD}",
                               timeout_s=g.RD_LINE_TIMEOUT, quiet=True)

                print(f"  capturing {n} samples (~{n / 10:.0f}s @10SPS)...")
                values = g.capture_samples(ser, n)
                g.save_csv(values, dut, mode)
                self._write_excel_retry(values, dut, mode, excel)

                if mode == "internal":
                    std, enob = summarize_internal(values)
                    if std is not None:
                        line = (f"Internal Short: STD = {std:,.2f}, "
                                f"ENOB = {enob:.2f} bit" if enob is not None
                                else f"Internal Short: STD = {std:,.2f}")
                        print(f"  >> {line}")
                        summary.append(line)
                else:
                    avg, vin_c, vin_a, acc = summarize_channel_a(
                        values, p["ainp"], p["ainn"])
                    if avg is not None:
                        line = (f"Channel A: 평균 = {avg:,.0f}, "
                                f"Vin(calc) = {vin_c * 1000:.3f} mV / "
                                f"{vin_a * 1000:.1f} mV")
                        if acc is not None:
                            line += f", 정확도 = {acc * 100:.3f} %"
                        print(f"  >> {line}")
                        summary.append(line)

        # 블록 메타: 실측 VDD/AVDD + 캘 반영 실지령 자극 (레거시 잔존값 대체)
        self._write_meta_retry(excel, dut, vdd_v, vref_v, p["ainp"], p["ainn"])
        return dut, summary


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
