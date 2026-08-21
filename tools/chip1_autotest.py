"""
CHIP1 24bit ADC characterization automation — v4 펌웨어(내장 DAC 통합) 이식판.

구판(old test/chip1_autotest.py)과의 차이:
  - 보드 1개 (Nucleo-H533RE 통합) — DAC 보드/칩선택 메뉴/보드 식별 로직 삭제
  - 깨끗한 CLI: NUL/ANSI 정리 불필요, 모든 명령이 'OK'/'ERR ...'로 종료
  - rr 응답 = '0x??' 줄 + 'OK' 줄
  - 내장 DAC: 부팅(전원 인가)마다 cal 계수 재입력 필수, cal 후 set 재실행 필수
  - wr 검증은 rd 2로 프레임을 흘린 뒤 rr (설정은 다음 변환 사이클부터 적용)
  - 캡처 전 세틀링: rd 5 폐기 (+ 필요 시 --settle-sec)

채널/전압 매핑 (팀 확정 표준 자극):
  ch1(PA4)→AINP=1.52V, ch2(PA5)→AINN=1.48V → 차동 +40mV → 평균 양수(약 +6.9M)
  구 엑셀의 Channel A 평균(+6.9M)과 부호까지 일치. 참고: 2026-07-20 E2E 검증은
  극성 반대(-40mV)로 수행됐음 (SPEC.md §5) — 그때 평균이 -6.871M이었으므로
  본측정에서는 +6.87M 근처가 나와야 정상.

Usage:
    pip install pyserial openpyxl
    python tools/chip1_autotest.py --dut 1
    python tools/chip1_autotest.py --port COM5 --dut 3 --settle-sec 60

엑셀: 기존 리포트(old test/26.7.14.adc.xlsx)는 건드리지 않는다. 첫 실행 시
같은 구조의 검증용 사본(CHIP1_ADC_validation.xlsx)을 만들고(데이터 영역만
비움, 수식/통계 행 유지) 이후 그 사본의 DUT#n 블록에 기입한다.
"""

import argparse
import json
import re
import shutil
import statistics
import sys
import time
from datetime import datetime
from pathlib import Path

# 내장 라이브러리 폴백: pip 설치가 없어도 tools/_vendor의 pyserial/openpyxl 사용
# (시스템 설치본이 있으면 그쪽 우선 — append라서)
_VENDOR = str(Path(__file__).resolve().parent / "_vendor")
if Path(_VENDOR).exists() and _VENDOR not in sys.path:
    sys.path.append(_VENDOR)

import serial
from serial.tools import list_ports

# ----------------------------- configuration -----------------------------

BAUD = 115200
LINE_END = "\r"            # 펌웨어는 \r/\n/\r\n 모두 허용
CMD_TIMEOUT = 5.0          # 일반 명령 OK 대기 (dac init 포함 여유)
RD_LINE_TIMEOUT = 5.0      # rd 샘플 줄 간 대기 (DRDY 타임아웃 500ms + 여유)

# 표준 자극 (팀 확정): AINP=1.52V, AINN=1.48V → 차동 +40mV (구 데이터와 부호 일치)
AINP_UV = 1520000          # ch1 (PA4) → AINP
AINN_UV = 1480000          # ch2 (PA5) → AINN
# 기본 보정계수 (SPEC.md §4, 2026-07-20 실측) — dac_cal.json이 있으면 그쪽 우선
DAC_CAL = {1: (0, -10137), 2: (0, -10529)}   # ch: (offset_uV, gain_ppm)

N_SAMPLES = 1024
SETTLE_RD = 5              # 본측정 전 폐기할 샘플 수 (rd 5)

PROJECT_ROOT = Path(__file__).resolve().parent.parent


def _resolve(primary: Path, *fallbacks: Path) -> Path:
    """새 폴더 구조 우선, 구버전 위치 폴백 (마이그레이션 전 호환)."""
    if primary.exists():
        return primary
    for f in fallbacks:
        if f.exists():
            return f
    return primary


# ---- 산출물/설정 폴더 구조 (유일한 경로 관리 지점 — GUI/capture도 이걸 사용) ----
CONFIG_DIR = PROJECT_ROOT / "config"           # 보드 고유 설정 (캘 계수)
RESULTS_DIR = PROJECT_ROOT / "results"
SWEEPS_DIR = RESULTS_DIR / "sweeps"            # 스윕 1회 = 폴더 1개 (타임스탬프)
MANUAL_DIR = RESULTS_DIR / "manual"            # 단발 캡처 (capture.py / Run test)

TEMPLATE_XLSX = _resolve(PROJECT_ROOT / "old test" / "26.7.14.adc.xlsx",
                         PROJECT_ROOT / "docs" / "26.7.14.adc_legacy_DAC1220.xlsx")
EXCEL_PATH = _resolve(RESULTS_DIR / "CHIP1_ADC_validation.xlsx",
                      PROJECT_ROOT / "CHIP1_ADC_validation.xlsx")
CSV_DIR = MANUAL_DIR                           # save_csv 기본 위치
CAL_FILE = _resolve(CONFIG_DIR / "dac_cal.json",
                    PROJECT_ROOT / "dac_cal.json",
                    Path(__file__).resolve().parent / "dac_cal.json")
MEAS_CAL_FILE = _resolve(CONFIG_DIR / "meas_cal.json",
                         PROJECT_ROOT / "meas_cal.json")
# 이 PC가 본 보드들의 기록 (UID -> 캘 이력). 진실은 보드 플래시, 이건 이력/추적용.
BOARD_REGISTRY = CONFIG_DIR / "board_registry.json"

SHEET_NAME = None          # None = active sheet
HEADER_ROW = 5             # '2026-07-15 DUT#n' 병합 헤더 행
MODE_ROW = 10              # 'Internal Short' / 'Channel A' 행
DATA_START_ROW = 22        # 첫 샘플 행 (B22=0 ... B1045=1023)

# 블록 메타데이터 행 (실측/실지령 자동 기입 대상 — 레거시 수기 값 대체)
META_ROW_VDD = 6           # 'Input Voltage (V)'  <- meas VDD 실측
META_ROW_AVDD = 7          # 'AVDD (V)'           <- meas VREF(VDDA) 실측
META_ROW_PGA = 8           # 'PGA (V/V)'
META_ROW_REG = 9           # '(0x03 0x02)' 레지스터 표기
META_ROW_ACTUAL_VIN = 16   # 'Actual Vin (V)'     <- 캘 보정 반영 실지령 차동
META_ROW_VIN_MV = 17       # 'DAC SET: VIN(mV)'   <- r18/r19 수식의 입력

# ----------------------------- serial layer -----------------------------


class CliError(RuntimeError):
    pass


def find_port() -> str:
    """ST-Link VCP(VID 0x0483) 단일 포트 자동 탐지."""
    cands = [p.device for p in list_ports.comports() if p.vid == 0x0483]
    if len(cands) == 1:
        return cands[0]
    if not cands:
        sys.exit("ST-Link VCP 포트를 찾지 못함. USB 연결 확인 또는 --port 지정.")
    sys.exit(f"ST-Link 포트가 여러 개: {cands}. --port로 지정하세요.")


def resync(ser: serial.Serial, tries: int = 3) -> bool:
    """세션 시작 재동기화: 이전 세션이 rd 스트리밍 중에 끊겼으면 샘플 줄이
    계속 흘러들어와 다음 명령 응답으로 오독된다 (실사고: id가 Short 샘플 값
    패턴으로 응답됨). ESC로 진행 중 rd를 중단시키고, 버퍼를 비운 뒤,
    빈 줄 -> 프롬프트('> ')만 오는 깨끗한 상태를 확인한다."""
    ser.write(b"\x1b")                 # ESC: 진행 중 rd면 'ERR aborted'로 중단
    ser.flush()
    time.sleep(0.6)                    # 중단 + 전송 중이던 잔여 줄 배출 대기
    for _ in range(tries):
        ser.reset_input_buffer()
        ser.write(LINE_END.encode())   # 빈 줄: 펌웨어는 무시하고 프롬프트만 출력
        ser.flush()
        time.sleep(0.3)
        resp = ser.read(ser.in_waiting or 1).decode(errors="replace")
        # 깨끗함 = 프롬프트가 보이고 샘플 스트림(큰 숫자 줄)이 없음
        if ">" in resp and not re.search(r"-?\d{4,}", resp):
            return True
        time.sleep(0.3)
    return False


# 칩 인터페이스: 'spi' = CHIP1(2선 펄스, 기본) / 'i2c' = CHIP1A(I2C 0x2A).
# GUI 셀렉션/CLI --iface가 set_iface()로 설정하면 open_port가 보드에 반영한다.
IFACE = "spi"


def set_iface(iface: str):
    global IFACE
    IFACE = "i2c" if str(iface).strip().lower() == "i2c" else "spi"


def open_port(port_name: str) -> serial.Serial:
    ser = serial.Serial(port_name, BAUD, timeout=0.2)
    time.sleep(0.3)
    resync(ser)                        # 이전 세션 잔여 상태 정리 (항상 수행)
    # 인터페이스 선택 반영. 이전 세션이 i2c로 두고 끊겼을 수 있어 spi도 명시
    # 전송 (iface 명령이 없는 구펌웨어면 spi에 한해 조용히 무시 — 부팅 기본값
    # 이 spi라 동작 동일. i2c 요청인데 구펌웨어면 에러를 그대로 낸다).
    try:
        send_cmd(ser, f"iface {IFACE}", quiet=True,
                 timeout_s=3.0)        # spi 복귀는 펌웨어가 350ms 웨이크 대기
    except CliError:
        if IFACE != "spi":
            raise CliError("이 펌웨어에 iface 명령이 없습니다 — I2C(CHIP1A) "
                           "모드는 최신 펌웨어 플래시 후 사용 가능")
    return ser


def _read_line(ser: serial.Serial, timeout_s: float) -> str:
    """한 줄 수신. flat 프롬프트 '> ' 잔여물은 제거."""
    deadline = time.time() + timeout_s
    buf = b""
    while time.time() < deadline:
        ch = ser.read(1)
        if not ch:
            continue
        if ch == b"\n":
            line = buf.decode(errors="replace").strip()
            buf = b""
            while line.startswith("> "):
                line = line[2:].strip()
            if line:
                return line
            deadline = time.time() + timeout_s
            continue
        if ch != b"\r":
            buf += ch
    raise CliError(f"응답 타임아웃 ({timeout_s:.0f}s)")


def send_cmd(ser: serial.Serial, cmd: str, timeout_s: float = CMD_TIMEOUT,
             on_line=None, quiet: bool = False) -> list[str]:
    """명령 전송 → 'OK'까지의 페이로드 줄 리스트 반환. 'ERR ...'이면 예외.

    on_line 콜백을 주면 페이로드 줄마다 호출 (rd 스트리밍용).
    """
    ser.reset_input_buffer()
    ser.write((cmd + LINE_END).encode())
    ser.flush()
    if not quiet:
        print(f"  > {cmd}")
    payload: list[str] = []
    while True:
        line = _read_line(ser, timeout_s)
        if line == cmd:            # 에코
            continue
        if line == "OK":
            return payload
        if line.startswith("ERR"):
            raise CliError(f"'{cmd}' 실패: {line}")
        payload.append(line)
        if on_line:
            on_line(line)


# ----------------------------- device ops -----------------------------


def check_id(ser: serial.Serial):
    """id == 0x9210 이면 통신+배선+전원 정상.
    불일치 시 즉시 실패하지 않고 1회 재동기화 후 재시도 (잔여 스트림 오독 대비)."""
    last = None
    for attempt in (1, 2):
        try:
            payload = send_cmd(ser, "id")
        except CliError as e:
            payload, last = None, str(e)
        if payload and payload[0].lower() == "0x9210":
            print("    chip ID OK (0x9210)")
            return
        last = payload[0] if payload else last
        if attempt == 1:
            print("  !! 이전 세션 잔여 데이터 감지 — 재동기화 시도")
            resync(ser)
        else:
            raise CliError(f"칩 ID 불일치 (재동기화 후에도 {last!r}, 기대 0x9210). "
                           "배선/전원 확인.")


def load_dac_cal() -> dict[int, tuple[int, int]]:
    """채널별 (offset_uV, gain_ppm). dac_cal.json 우선, 없으면 기본값(DAC_CAL)."""
    cal = dict(DAC_CAL)
    if CAL_FILE.exists():
        try:
            raw = json.loads(CAL_FILE.read_text(encoding="utf-8"))
            for ch in (1, 2):
                e = raw.get(str(ch))
                if e is not None:
                    cal[ch] = (int(e["offset_uv"]), int(e["gain_ppm"]))
            print(f"  cal 계수: {CAL_FILE.name} 사용 {cal}")
        except (ValueError, KeyError, OSError, TypeError) as e:
            print(f"  !! {CAL_FILE.name} 읽기 실패({e}) — 기본 계수 사용 {cal}")
    else:
        print(f"  cal 계수: 기본값 사용 {cal} ({CAL_FILE.name} 없음)")
    return cal


def save_dac_cal(ch: int, offset_uv: int, gain_ppm: int, extra: dict | None = None):
    """2점 캘리브레이션 결과를 dac_cal.json에 채널별로 저장 (GUI 위저드가 호출)."""
    data = {}
    if CAL_FILE.exists():
        try:
            data = json.loads(CAL_FILE.read_text(encoding="utf-8"))
        except (ValueError, OSError):
            data = {}
    entry = {"offset_uv": int(offset_uv), "gain_ppm": int(gain_ppm),
             "date": datetime.now().isoformat(timespec="seconds")}
    if extra:
        entry.update(extra)
    data[str(ch)] = entry
    CAL_FILE.parent.mkdir(parents=True, exist_ok=True)
    CAL_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False),
                        encoding="utf-8")


# ---- meas(내장 ADC) 2점 캘리브레이션 — PC측 보정 (meas_cal.json) ----
# 채널 키: "vdd"(A3) / "vref"(A4). 적용식은 DAC과 동형:
#   corrected_uV = (raw_uV + offset_uv) * (1 + gain_ppm/1e6)

def load_meas_cal() -> dict:
    """{'vdd': (offset_uv, gain_ppm), 'vref': (...)} — 없으면 (0,0) = 무보정."""
    cal = {"vdd": (0, 0), "vref": (0, 0)}
    if MEAS_CAL_FILE.exists():
        try:
            raw = json.loads(MEAS_CAL_FILE.read_text(encoding="utf-8"))
            for ch in ("vdd", "vref"):
                e = raw.get(ch)
                if e is not None:
                    cal[ch] = (int(e["offset_uv"]), int(e["gain_ppm"]))
        except (ValueError, KeyError, OSError, TypeError) as e:
            print(f"  !! {MEAS_CAL_FILE.name} 읽기 실패({e}) — meas 무보정 사용")
    return cal


def save_meas_cal(ch: str, offset_uv: int, gain_ppm: int,
                  extra: dict | None = None):
    data = {}
    if MEAS_CAL_FILE.exists():
        try:
            data = json.loads(MEAS_CAL_FILE.read_text(encoding="utf-8"))
        except (ValueError, OSError):
            data = {}
    entry = {"offset_uv": int(offset_uv), "gain_ppm": int(gain_ppm),
             "date": datetime.now().isoformat(timespec="seconds")}
    if extra:
        entry.update(extra)
    data[ch] = entry
    MEAS_CAL_FILE.parent.mkdir(parents=True, exist_ok=True)
    MEAS_CAL_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False),
                             encoding="utf-8")


def read_uid(ser) -> str | None:
    """'uid' 명령 → 'UID=XXXXXXXX-XXXXXXXX-XXXXXXXX' 문자열. 구펌웨어면 None."""
    try:
        payload = send_cmd(ser, "uid")
    except CliError:
        return None
    for line in payload:
        m = re.search(r"UID=([0-9A-Fa-f]{8}-[0-9A-Fa-f]{8}-[0-9A-Fa-f]{8})", line)
        if m:
            return m.group(1)
    return None


def board_status(ser) -> dict:
    """연결된 보드 상태 조회: UID + 채널별 플래시 캘 상태 + 현재 적용 계수 +
    PC 레지스트리 기록.
    반환: {'uid', 'flash_ch': {1:bool,2:bool}, 'flash_cal': bool(양채널 완캘),
          'cal': {1:(off_uv,ppm), 2:(off_uv,ppm)}, 'known', 'record'}"""
    uid = read_uid(ser)
    flash_ch = {1: False, 2: False}
    cal = {1: (0, 0), 2: (0, 0)}
    try:
        lines = send_cmd(ser, "dac cal show")
        flash_ch = parse_flash_state(lines)
        for line in lines:
            m = re.match(r"ch([12])\s+offset_uV=(-?\d+)\s+gain_ppm=(-?\d+)", line)
            if m:
                cal[int(m.group(1))] = (int(m.group(2)), int(m.group(3)))
    except CliError:
        pass
    reg = _load_registry()
    rec = reg.get(uid) if uid else None
    return {"uid": uid, "flash_ch": flash_ch,
            "flash_cal": flash_ch[1] and flash_ch[2], "cal": cal,
            "known": rec is not None, "record": rec}


def _load_registry() -> dict:
    if BOARD_REGISTRY.exists():
        try:
            return json.loads(BOARD_REGISTRY.read_text(encoding="utf-8"))
        except (ValueError, OSError):
            return {}
    return {}


def registry_note(uid: str, **fields):
    """이 PC의 보드 레지스트리에 UID 기록/갱신 (last_seen 자동)."""
    if not uid:
        return
    reg = _load_registry()
    rec = reg.get(uid, {})
    rec.update(fields)
    rec["last_seen"] = datetime.now().isoformat(timespec="seconds")
    reg[uid] = rec
    BOARD_REGISTRY.parent.mkdir(parents=True, exist_ok=True)
    BOARD_REGISTRY.write_text(json.dumps(reg, indent=2, ensure_ascii=False),
                              encoding="utf-8")


def parse_meas(payload_lines) -> tuple[float | None, float | None]:
    """'VDD=3.334V VREF=3.023V' → (3.334, 3.023)."""
    for line in payload_lines:
        m = re.search(r"VDD=([\d.]+)V\s+VREF=([\d.]+)V", line)
        if m:
            return float(m.group(1)), float(m.group(2))
    return None, None


def meas_selfcal(ser) -> dict:
    """'meas cal' 실행: 펌웨어가 ADCAL(오프셋) 재실행 + VREFINT(게인) 재기준.
    반환: {'calfactor': int, 'vrefint_raw': int, 'vdda_v': float}"""
    payload = send_cmd(ser, "meas cal")
    for line in payload:
        m = re.search(r"CALFACTOR=(\d+)\s+VREFINT_RAW=(\d+)\s+VDDA=([\d.]+)V", line)
        if m:
            return {"calfactor": int(m.group(1)),
                    "vrefint_raw": int(m.group(2)),
                    "vdda_v": float(m.group(3))}
    raise CliError(f"meas cal 응답 파싱 실패: {payload!r}")


def save_selfcal_record(info: dict, verify_err_mv=None, verify=None):
    """셀프캘 이력을 meas_cal.json의 'selfcal' 항목에 기록 (계수 적용과 무관한
    기록용 — 실제 보정은 ADC 하드웨어 내부에서 이미 적용됨)."""
    entry = dict(info)
    entry["date"] = datetime.now().isoformat(timespec="seconds")
    if verify_err_mv is not None:
        entry["verify_err_mv"] = verify_err_mv
        entry["verify"] = verify
    data = {}
    if MEAS_CAL_FILE.exists():
        try:
            data = json.loads(MEAS_CAL_FILE.read_text(encoding="utf-8"))
        except (ValueError, OSError):
            data = {}
    data["selfcal"] = entry
    MEAS_CAL_FILE.parent.mkdir(parents=True, exist_ok=True)
    MEAS_CAL_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False),
                             encoding="utf-8")


def apply_meas_cal(vdd_v: float | None, vref_v: float | None):
    """meas 파싱값(V)에 meas_cal.json 보정 적용. None은 그대로 통과."""
    cal = load_meas_cal()

    def corr(v, key):
        if v is None:
            return None
        off, ppm = cal[key]
        return round((v * 1e6 + off) * (1.0 + ppm / 1e6) / 1e6, 4)

    return corr(vdd_v, "vdd"), corr(vref_v, "vref")


def parse_flash_state(lines) -> dict:
    """'flash: ch1=saved ch2=none' → {1: True, 2: False}.
    구형 단일 표기('flash: saved'/'none')도 하위 호환 수용."""
    st = {1: False, 2: False}
    for line in lines:
        low = line.lower()
        if not low.startswith("flash:"):
            continue
        m = re.search(r"ch1=(\w+)\s+ch2=(\w+)", low)
        if m:
            st[1] = m.group(1) == "saved"
            st[2] = m.group(2) == "saved"
        elif "saved" in low:
            st[1] = st[2] = True
    return st


def dac_ready(ser: serial.Serial) -> str:
    """dac init + 캘 준비. **보드 플래시 캘이 진실** — 저장된 채널은 그 값을
    신뢰하고 json 주입 생략 (stale json이 보드 캘을 덮는 사고 방지). 저장이
    없는 채널만 json/기본 계수 폴백. 반환: 'flash' | 'partial' | 'json'."""
    send_cmd(ser, "dac init")            # 플래시 캘(valid 채널) 자동 로드됨
    try:
        st = parse_flash_state(send_cmd(ser, "dac cal show"))
    except CliError:
        st = {1: False, 2: False}        # 구펌웨어 -> 전량 json 폴백
    missing = [ch for ch in (1, 2) if not st[ch]]
    if not missing:
        print("  cal: 보드 플래시 저장값 사용 (양 채널 — json 주입 생략)")
        return "flash"
    jcal = load_dac_cal()
    for ch in missing:
        off, ppm = jcal[ch]
        send_cmd(ser, f"dac cal {ch} {off} {ppm}")
    if len(missing) == 2:
        print("  cal: 보드 저장 없음 -> 전 채널 json/기본 계수 (GUI 캘+save 권장)")
        return "json"
    print(f"  cal: ch{missing[0]}만 json 폴백 (보드 미캘 채널 — GUI 캘 권장)")
    return "partial"


def setup_dac(ser: serial.Serial):
    """DAC 준비 + 표준 자극 출력. 순서: init/cal(dac_ready) → set."""
    print("\n[internal DAC setup]")
    dac_ready(ser)
    send_cmd(ser, f"dac set 1 {AINP_UV}")   # ch1(PA4) → AINP
    send_cmd(ser, f"dac set 2 {AINN_UV}")   # ch2(PA5) → AINN


def write_verified(ser: serial.Serial, addr: str, val: str):
    """wr → rd 2(설정 반영 프레임 흘리기) → rr 검증. 1회 재시도."""
    for attempt in (1, 2):
        send_cmd(ser, f"wr {addr} {val}")
        send_cmd(ser, "rd 2", timeout_s=RD_LINE_TIMEOUT, quiet=True)  # flush frames
        payload = send_cmd(ser, f"rr {addr}")
        got = int(payload[0], 16) if payload else -1
        if got == int(val, 16):
            return
        print(f"  !! rr {addr} = {payload[0] if payload else '?'} "
              f"(기대 {val}), 재시도 {attempt}/2")
    sys.exit(f"레지스터 {addr} 검증 2회 실패. 중단.")


def capture_samples(ser: serial.Serial, n: int) -> list[int]:
    """rd <n> 스트리밍 수집. 샘플당 한 줄(부호 있는 10진), 마지막 OK."""
    values: list[int] = []
    t0 = time.time()

    def on_line(line: str):
        try:
            values.append(int(line, 10))
        except ValueError:
            pass                       # 배너 등 무관한 줄 무시
        if len(values) % 100 == 0 and values:
            print(f"    {len(values)}/{n}")

    # 10SPS 기준 n/10초 + 여유. 줄 간 타임아웃은 RD_LINE_TIMEOUT이 지킨다.
    send_cmd(ser, f"rd {n}", timeout_s=RD_LINE_TIMEOUT, on_line=on_line, quiet=True)
    print(f"  capture complete: {len(values)}개 in {time.time() - t0:.1f}s")
    if len(values) != n:
        sys.exit(f"샘플 수 불일치: {len(values)}/{n}. 중단.")
    return values


# ----------------------------- excel / csv -----------------------------


def save_csv(values: list[int], dut: int, mode: str,
             out_dir: Path | None = None) -> Path:
    """샘플 CSV 저장. 기본은 results/manual/, 스윕은 런 폴더의 captures/를 지정."""
    d = out_dir or CSV_DIR
    d.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = d / f"dut{dut}_{mode}_{ts}.csv"
    with open(path, "w", newline="") as f:
        f.write("index,code\n")
        for i, v in enumerate(values):
            f.write(f"{i},{v}\n")
    mean = statistics.fmean(values)
    stdev = statistics.stdev(values) if len(values) > 1 else 0.0
    print(f"  saved {path.parent.name}/{path.name}  (mean {mean:,.1f}, stdev {stdev:,.1f})")
    return path


def _i2c_sheet_name(wb) -> str:
    """I2C 결과 시트 이름 결정.

    핵심 규칙: **이미 존재하는 '*_I2C' 시트가 있으면 그걸 재사용** — 활성
    탭이 무엇이든 중복 시트를 만들지 않는다 (실사고: 활성=260410_SPI 상태에서
    260807_I2C가 있는데 260410_SPI_I2C를 새로 만듦). 없을 때만 활성 시트
    이름 기반으로 짓되, '_SPI' 접미는 떼고 '_I2C'를 붙인다."""
    active = wb.active.title
    if active.endswith("_I2C"):
        return active
    existing = [s for s in wb.sheetnames if s.endswith("_I2C")]
    if existing:
        return existing[0]
    base = active[:-4] if active.endswith("_SPI") else active
    return base + "_I2C"


def _spi_sheet_name(wb) -> str:
    """SPI 결과 시트 이름 결정 (_i2c_sheet_name과 대칭): 기존 '*_SPI' 시트
    재사용 우선, 없으면 일반 시트 이름 + '_SPI'."""
    active = wb.active.title
    if active.endswith("_SPI"):
        return active
    existing = [s for s in wb.sheetnames if s.endswith("_SPI")]
    if existing:
        return existing[0]
    base = active[:-4] if active.endswith("_I2C") else active
    if base.startswith("SW_"):
        for s in wb.sheetnames:
            if not s.startswith("SW_") and not s.endswith("_I2C"):
                base = s[:-4] if s.endswith("_SPI") else s
                break
    return base + "_SPI"


def _spi_sheet(wb):
    """SPI 결과 시트: '*_SPI' 시트 우선 (ensure_iface_sheet가 개명해 둠).
    아직 개명 전 워크북 폴백: 활성(비 I2C) → base → 비스윕 일반 시트."""
    name = _spi_sheet_name(wb)
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.active
    if not ws.title.endswith("_I2C"):
        return ws
    base = ws.title[:-4]
    if base in wb.sheetnames:
        return wb[base]
    for s in wb.sheetnames:
        if not s.endswith("_I2C") and not s.startswith("SW_"):
            return wb[s]
    return ws


def _get_sheet(wb):
    """대상 시트 라우팅 — 칩 인터페이스별 결과 분리의 단일 지점.

    - SHEET_NAME 명시(스윕 시트/--sheet) → 그대로 (스윕은 시트명 자체에
      I2C 태그가 붙음, create_sweep_sheet 참조)
    - 기본(상온): SPI = 일반 시트(_I2C 회피) / I2C = 기존 '*_I2C' 시트 재사용
      (ensure_iface_sheet가 미리 생성해 둠 — 없으면 명확히 실패)
    """
    if SHEET_NAME:
        return wb[SHEET_NAME]
    if IFACE == "i2c":
        name = _i2c_sheet_name(wb)
        if name not in wb.sheetnames:
            raise CliError(f"I2C 결과 시트({name})가 없음 — "
                           "테스트 시작 경로에서 ensure_iface_sheet() 필요")
        return wb[name]
    return _spi_sheet(wb)


def ensure_iface_sheet(excel_path: Path):
    """상온 결과 시트를 인터페이스별 이름으로 보장 (테스트 시작 시 호출).

    - i2c: '*_I2C' 시트 재사용, 없으면 SPI 시트 레이아웃 복제로 생성
    - spi: '*_SPI' 시트 재사용, 없으면 기본 시트를 '<이름>_SPI'로 **개명**
      (기존 데이터 보존 — 복제/신규 생성 아님)"""
    if SHEET_NAME:
        return None
    from openpyxl import load_workbook

    if IFACE != "i2c":
        wb = load_workbook(excel_path)
        name = _spi_sheet_name(wb)
        if name in wb.sheetnames:
            wb.close()
            return name
        ws = wb.active
        if ws.title.startswith("SW_") or ws.title.endswith("_I2C"):
            for s in wb.sheetnames:
                if not s.startswith("SW_") and not s.endswith("_I2C"):
                    ws = wb[s]
                    break
            else:
                wb.close()
                return None            # 개명할 일반 시트가 없음 (비정상 구성)
        old = ws.title
        ws.title = old + "_SPI"
        wb.save(excel_path)
        wb.close()
        print(f"SPI 결과 시트 개명: {old} -> {old}_SPI (데이터 보존)")
        return old + "_SPI"

    wb = load_workbook(excel_path)
    name = _i2c_sheet_name(wb)          # 기존 *_I2C 시트 재사용 (중복 방지)
    if name in wb.sheetnames:
        wb.close()
        return name

    base = _spi_sheet(wb)               # 레이아웃 원본 = SPI 시트
    _clone_layout_sheet(wb, base, name)
    wb.save(excel_path)
    print(f"I2C 전용 시트 생성: {name} (기본 시트 레이아웃 복제, 데이터 비움)")
    return name


def _clone_layout_sheet(wb, base, title):
    """base 시트를 레이아웃만 복제 (샘플/메타 데이터 비움, 수식 보존)."""
    ws = wb.copy_worksheet(base)
    ws.title = title

    def is_formula(cell):
        return cell.data_type == "f" or (isinstance(cell.value, str)
                                         and cell.value.startswith("="))

    for _dut, c in _find_dut_blocks(ws):
        for col in (c, c + 1):
            for r in range(DATA_START_ROW, DATA_START_ROW + N_SAMPLES):
                if ws.cell(row=r, column=col).value is not None:
                    ws.cell(row=r, column=col).value = None
            for r in (META_ROW_VDD, META_ROW_AVDD, META_ROW_PGA, META_ROW_REG):
                cell = ws.cell(row=r, column=col)
                if not is_formula(cell):
                    cell.value = None
        ws.cell(row=META_ROW_ACTUAL_VIN, column=c + 1).value = None
        c17 = ws.cell(row=META_ROW_VIN_MV, column=c + 1)
        if not is_formula(c17):
            c17.value = None
        hcell = ws.cell(row=HEADER_ROW, column=c)
        if isinstance(hcell.value, str):
            hcell.value = re.sub(r"\s*@.*$", "", hcell.value)
    _strip_header_dates(ws)     # 복제 시트도 날짜 비움 (기록 시 실날짜 기입)
    return ws


def ensure_named_sheet(excel_path: Path, name: str) -> str:
    """이름 지정 결과 시트 보장 — **새 테스트 그룹용** (2026-08-18).

    있으면 그대로 재사용, 없으면 기본(SPI) 시트 레이아웃을 복제해 데이터를
    비운 새 시트 생성. SHEET_NAME과 함께 쓰면 ADC 결과가 그 시트로 기록됨."""
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    if name in wb.sheetnames:
        wb.close()
        return name
    base = _spi_sheet(wb)
    _clone_layout_sheet(wb, base, name)
    wb.save(excel_path)
    wb.close()
    print(f"새 그룹 시트 생성: {name} (레이아웃 복제, 데이터 비움)")
    return name


def _find_dut_blocks(ws) -> list[tuple[int, int]]:
    """헤더 행에서 (dut번호, 앵커열) 목록."""
    pattern = re.compile(r"DUT\s*#\s*(\d+)")
    out = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if isinstance(v, str):
            m = pattern.search(v)
            if m:
                out.append((int(m.group(1)), c))
    return out


def _workbook_is_blank(excel_path: Path) -> bool:
    """사용자가 직접 만든 '빈 통합문서'인지 판정 — 모든 시트에 값이 하나도
    없으면 빈 것으로 본다 (엑셀 새 문서의 기본 Sheet1 등)."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(excel_path)
    except Exception:                     # noqa: BLE001 - 못 여는 파일은 보호
        return False
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        return False
        return True
    finally:
        wb.close()


def _workbook_has_layout(excel_path: Path) -> bool:
    """어느 시트든 DUT 블록 헤더가 있으면 레이아웃 보유로 판정."""
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    try:
        return any(_find_dut_blocks(ws) for ws in wb.worksheets)
    finally:
        wb.close()


def _strip_header_dates(ws):
    """블록 헤더의 박제 날짜 제거 — '2026-07-14 DUT#1' → 'DUT#1'.
    실제 날짜는 결과가 기록될 때 write_block_meta가 그날 날짜로 채움."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if isinstance(v, str) and re.match(r"\s*\d{4}-\d{2}-\d{2}", v):
            ws.cell(row=HEADER_ROW, column=c).value = \
                re.sub(r"^\s*\d{4}-\d{2}-\d{2}\s*", "", v, count=1)


def ensure_validation_workbook(excel_path: Path, template: Path,
                               sheet_name: str | None = None):
    """검증용 사본이 없으면 생성: 템플릿 복사 후 샘플 데이터 영역만 비움.
    수식/통계/헤더는 그대로 유지 → 새 데이터 입력 시 자동 재계산.

    2026-08-18 정비:
    - 사용자가 미리 만든 **빈 엑셀**(값 없는 새 통합문서)도 템플릿으로 시딩.
      값이 있는데 DUT 블록이 없는 파일은 데이터 보호를 위해 안내 후 중단.
    - 신규 생성/시딩 시 **블록 헤더의 박제 날짜 제거** (기록 시 실제 날짜로
      채워짐) + 상단 '날짜, HW Team' 줄 제거.
    - sheet_name 지정 시 템플릿 기본 시트를 그 이름으로 **개명** — 새 파일이
      곧바로 사용자가 정한 시트 하나로 시작 (별도 복제 시트 안 생김)."""
    if excel_path.exists():
        if _workbook_has_layout(excel_path):
            return
        if _workbook_is_blank(excel_path):
            print(f"빈 통합문서 감지: {excel_path.name} — 템플릿 레이아웃으로 시딩")
            excel_path.unlink()           # 아래 신규 생성 경로로 진행
        else:
            raise CliError(
                f"{excel_path.name}에 DUT 블록 레이아웃이 없습니다. "
                "빈 파일이면 내용을 비우고 재실행(자동 시딩), 데이터가 있는 "
                "파일이면 template\\report_template.xlsx의 시트를 복사해 넣으세요")
    if not template.exists():
        sys.exit(f"템플릿 없음: {template}")
    from openpyxl import load_workbook
    shutil.copyfile(template, excel_path)
    wb = load_workbook(excel_path)
    ws = _get_sheet(wb)
    blocks = _find_dut_blocks(ws)
    for _dut, c in blocks:
        for col in (c, c + 1):                    # Internal Short, Channel A
            for r in range(DATA_START_ROW, DATA_START_ROW + N_SAMPLES):
                ws.cell(row=r, column=col).value = None
    # 레거시 박제값 정리: 블록 헤더 날짜 + 상단 '2026-07-14, HW Team' 줄
    _strip_header_dates(ws)
    for r in range(1, HEADER_ROW):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and re.match(r"\s*\d{4}-\d{2}-\d{2}", v):
                ws.cell(row=r, column=c).value = None
    if sheet_name:
        ws.title = sheet_name             # 새 파일 = 지정 시트 하나로 시작
    wb.save(excel_path)
    print(f"검증용 워크북 생성: {excel_path.name} "
          f"(원본 {template.name} 구조, {len(blocks)}개 DUT 블록 비움"
          + (f", 시트명 '{sheet_name}'" if sheet_name else "") + ")")
    print("  * openpyxl 저장 시 차트/일부 서식이 소실될 수 있음. 원본은 그대로니 확인만.")


def locate_dut_columns(ws, dut: int) -> tuple[int, int]:
    pattern = re.compile(rf"DUT\s*#\s*{dut}\b")
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if isinstance(v, str) and pattern.search(v):
            mode_int = str(ws.cell(row=MODE_ROW, column=c).value or "")
            mode_cha = str(ws.cell(row=MODE_ROW, column=c + 1).value or "")
            if "Internal" not in mode_int or "Channel" not in mode_cha:
                sys.exit(f"DUT#{dut} 헤더(열 {c})는 찾았으나 모드 행 불일치 "
                         f"({mode_int!r} / {mode_cha!r}). 레이아웃 확인.")
            return c, c + 1
    sys.exit(f"헤더 행 {HEADER_ROW}에서 'DUT#{dut}'를 찾지 못함.")


def find_next_empty_dut(ws) -> int | None:
    for dut, c in sorted(_find_dut_blocks(ws)):
        if ws.cell(row=DATA_START_ROW, column=c).value is None:
            return dut
    return None


# 주의: '실제 출력 추정값' = 지령값 그대로다. 캘리브레이션이 펌웨어 안에서
# (uv+off)(1+ppm)을 적용해 물리 출력을 지령에 일치시키므로 (검증 ±1mV),
# 지령에 보정식을 또 곱하면 안 됨 (2026-07-24 자가 수정).


def write_block_meta(excel_path: Path, dut: int,
                     vdd_v: float | None, vref_v: float | None,
                     ainp_uv: int, ainn_uv: int,
                     pga: int = 64, reg_note: str = "(0x03 0x02)"):
    """블록 메타 행에 이번 런의 실측/실지령 기입 (레거시 잔존값 대체).

    - r6/r7: meas 실측 VDD/AVDD (없으면 건드리지 않음 — 스테일 방지 위해 빈칸)
    - r16: Actual Vin = 캘 보정 반영 지령 차동 (수식 '=AINP-AINN' 형태로 출처 보존)
    - r17: VIN(mV) — r18/r19의 AINN/AINP 수식이 이 값에서 파생됨
    수식 셀(r15/r18/r19/r20 등)은 절대 덮지 않는다.
    """
    from openpyxl import load_workbook

    wb = load_workbook(excel_path)
    ws = _get_sheet(wb)
    col_int, col_cha = locate_dut_columns(ws, dut)

    # 블록 안에서도 인터페이스가 보이게 레지스터 표기에 태그 (시트 분리와 별개)
    if IFACE == "i2c" and "I2C" not in reg_note:
        reg_note = (reg_note[:-1] + ", I2C)") if reg_note.endswith(")") \
                   else reg_note + " I2C"

    # 블록 헤더 날짜를 실제 테스트 날짜로 갱신 (레거시 템플릿의 박제 날짜
    # 대체). 형식 'YYYY-MM-DD DUT#n[ @온도 [태그]]' — 접미(@/[])는 보존.
    hcell = ws.cell(row=HEADER_ROW, column=col_int)
    hval = str(hcell.value or f"DUT#{dut}")
    today = datetime.now().strftime("%Y-%m-%d")
    if re.match(r"\s*\d{4}-\d{2}-\d{2}", hval):
        hval = re.sub(r"^\s*\d{4}-\d{2}-\d{2}", today, hval, count=1)
    else:
        hval = f"{today} {hval.strip()}"
    hcell.value = hval

    def put(row, col, value):
        cell = ws.cell(row=row, column=col)
        # 수식 보존: 문자열 '=' 수식뿐 아니라 배열수식(ArrayFormula 객체,
        # 예: r12 STDEV)도 data_type 'f'로 걸러냄 (2026-07-24 리뷰 보강)
        if cell.data_type == "f" or (isinstance(cell.value, str)
                                     and cell.value.startswith("=")):
            return          # (r16만 예외적으로 아래서 직접 교체)
        cell.value = value

    for col in (col_int, col_cha):
        put(META_ROW_VDD, col, vdd_v if vdd_v is not None else None)
        put(META_ROW_AVDD, col, vref_v if vref_v is not None else None)
        put(META_ROW_PGA, col, pga)
        put(META_ROW_REG, col, reg_note)

    ainp_v = round(ainp_uv / 1e6, 6)
    ainn_v = round(ainn_uv / 1e6, 6)
    # r16은 레거시가 수기 수식이라 명시적으로 교체 (출처가 보이는 수식 형태)
    ws.cell(row=META_ROW_ACTUAL_VIN, column=col_cha).value = f"={ainp_v}-{ainn_v}"
    put(META_ROW_VIN_MV, col_cha, round((ainp_uv - ainn_uv) / 1000.0, 3))

    wb.save(excel_path)


def write_excel(values: list[int], dut: int, mode: str, excel_path: Path):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)
    ws = _get_sheet(wb)
    col_int, col_cha = locate_dut_columns(ws, dut)
    col = col_int if mode == "internal" else col_cha

    existing = ws.cell(row=DATA_START_ROW, column=col).value
    if existing is not None:
        ans = input(f"  DUT#{dut} {mode} 열에 이미 데이터 있음 "
                    f"(행 {DATA_START_ROW} = {existing}). 덮어쓸까요? [y/N] ").strip().lower()
        if ans != "y":
            print("  엑셀 기입 생략 (CSV는 저장됨).")
            return
    for i, v in enumerate(values):
        ws.cell(row=DATA_START_ROW + i, column=col, value=v)
    wb.save(excel_path)
    print(f"  wrote {len(values)} samples -> {excel_path.name} "
          f"[{ws.title}!{get_column_letter(col)}{DATA_START_ROW}:"
          f"{get_column_letter(col)}{DATA_START_ROW + len(values) - 1}]")


# ----------------------------- test sequences -----------------------------


def run_mode(ser: serial.Serial, dut: int, mode: str, reg04: str,
             excel_path: Path, settle_sec: float):
    label = "Internal Short" if mode == "internal" else "Channel A"
    print(f"\n[{label}] DUT #{dut}")

    write_verified(ser, "0x03", "0x02")     # 10SPS, PGA x64
    write_verified(ser, "0x04", reg04)

    if settle_sec > 0:
        print(f"  settling wait {settle_sec:.0f}s (과도 드리프트 회피)")
        time.sleep(settle_sec)
    if SETTLE_RD > 0:
        print(f"  discarding {SETTLE_RD} settling samples")
        send_cmd(ser, f"rd {SETTLE_RD}", timeout_s=RD_LINE_TIMEOUT, quiet=True)

    print(f"  capturing {N_SAMPLES} samples (~{N_SAMPLES / 10:.0f}s @10SPS)...")
    values = capture_samples(ser, N_SAMPLES)
    save_csv(values, dut, mode)
    write_excel(values, dut, mode, excel_path)


def wait_for_port(poll: float = 1.0) -> str:
    print("  waiting for the board on USB", end="", flush=True)
    while True:
        cands = [p.device for p in list_ports.comports() if p.vid == 0x0483]
        if len(cands) == 1:
            print()
            time.sleep(1.5)            # 윈도우 열거 완료 대기
            return cands[0]
        print(".", end="", flush=True)
        time.sleep(poll)


def test_one_dut(port: str, dut: int, excel_path: Path, settle_sec: float):
    """칩 1개 전체 사이클: id 확인 → DAC 셋업 → meas → Short → ChA → 메타 기입."""
    with open_port(port) as ser:
        check_id(ser)
        setup_dac(ser)   # 전원 사이클로 cal이 소실됐을 수 있으니 매번 전체 재실행
        try:
            vdd_v, vref_v = apply_meas_cal(*parse_meas(send_cmd(ser, "meas")))
        except CliError:
            vdd_v = vref_v = None
        run_mode(ser, dut, "internal", "0x60", excel_path, settle_sec)
        run_mode(ser, dut, "channel_a", "0x00", excel_path, settle_sec)
    if excel_path.exists():
        write_block_meta(excel_path, dut, vdd_v, vref_v, AINP_UV, AINN_UV)
        print(f"  블록 메타 기입: VDD={vdd_v} AVDD={vref_v} "
              f"AINP/AINN={AINP_UV}/{AINN_UV}uV")


def main():
    global SHEET_NAME
    ap = argparse.ArgumentParser()
    ap.add_argument("--port", default=None, help="시리얼 포트 (기본: ST-Link 자동 탐지)")
    ap.add_argument("--dut", type=int, default=None,
                    help="시작 DUT 번호 (기본: 검증 워크북의 첫 빈 블록)")
    ap.add_argument("--excel", type=Path, default=EXCEL_PATH,
                    help=f"검증용 엑셀 경로 (기본: {EXCEL_PATH.name})")
    ap.add_argument("--sheet", default=None, help="시트 이름 (기본: active)")
    ap.add_argument("--settle-sec", type=float, default=0,
                    help="설정 후 캡처 전 추가 대기 초 (과도 드리프트 실측 시 60 권장)")
    ap.add_argument("--iface", choices=("spi", "i2c"), default="spi",
                    help="칩 인터페이스: spi=CHIP1(기본), i2c=CHIP1A")
    args = ap.parse_args()
    if args.sheet:
        SHEET_NAME = args.sheet
    set_iface(args.iface)

    ensure_validation_workbook(args.excel, TEMPLATE_XLSX)
    ensure_iface_sheet(args.excel)      # i2c면 '<시트>_I2C' 전용 시트 보장

    dut = args.dut
    if dut is None:
        from openpyxl import load_workbook
        wb = load_workbook(args.excel, read_only=True)
        dut = find_next_empty_dut(_get_sheet(wb))
        wb.close()
        if dut is None:
            sys.exit("빈 DUT 블록이 없음. --dut로 지정하세요.")
        print(f"자동 선택: 첫 빈 블록 DUT#{dut}")

    while True:
        port = args.port or wait_for_port()
        test_one_dut(port, dut, args.excel, args.settle_sec)

        print(f"\nDUT #{dut} complete.")
        ans = input("전원 분리, 칩 교체, 재연결 후 Enter (q = 종료): ").strip().lower()
        if ans == "q":
            break
        dut += 1
        args.port = None   # 재열거로 포트가 바뀔 수 있으니 다음 루프에서 재탐지

    print("\nAll done.")


if __name__ == "__main__":
    main()
